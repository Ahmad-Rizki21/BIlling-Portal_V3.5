import pandas as pd
from typing import List
import math
import openpyxl
from io import BytesIO
from typing import Optional
from datetime import datetime, timedelta, date, timezone
import io
import csv
import chardet  # Add this import for encoding detection
import json
import uuid
from typing import Union, TYPE_CHECKING

from fastapi import APIRouter, Depends, HTTPException, status, Request, Header, Query
from fastapi.responses import StreamingResponse, Response
from dateutil import parser
from dateutil.relativedelta import relativedelta

from sqlalchemy import func, or_, and_
from ..models.user import User as UserModel
from ..models.role import Role as RoleModel
from ..models.diskon import Diskon as DiskonModel
from ..websocket_manager import manager

# Import has_permission function
from ..auth import has_permission, get_current_active_user

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import joinedload, selectinload
from pydantic import ValidationError
import logging

from ..models.invoice import Invoice as InvoiceModel
from ..models.invoice_archive import InvoiceArchive as InvoiceArchiveModel
from ..models.langganan import Langganan as LanggananModel
from ..models.pelanggan import Pelanggan as PelangganModel
from ..schemas.invoice import (
    Invoice as InvoiceSchema,
    InvoiceGenerate,
    MarkAsPaidRequest,
)
from ..database import get_db

from sqlalchemy import select, func
from ..services import mikrotik_service
from ..config import settings
from ..services import xendit_service, mikrotik_service
from ..services.payment_callback_service import check_duplicate_callback, log_callback_processing
from ..services.rate_limiter import create_invoice_with_rate_limit, InvoicePriority
from ..utils.phone_utils import normalize_phone_for_xendit

# Import our logging utilities
from ..logging_utils import sanitize_log_data

logger = logging.getLogger("app.routers.invoice")


# Helper functions untuk safe date conversion
def safe_to_datetime(date_obj) -> datetime:
    """Convert date/datetime ke datetime dengan aman."""
    if date_obj is None:
        return datetime.now(timezone.utc)

    if isinstance(date_obj, datetime):
        return date_obj

    # Handle SQLAlchemy Date object atau Python date
    try:
        if hasattr(date_obj, "strftime"):
            return datetime.combine(date_obj, datetime.min.time())
        else:
            # Fallback untuk SQLAlchemy Date
            return datetime.combine(date_obj, datetime.min.time())
    except (AttributeError, TypeError):
        return datetime.now(timezone.utc)


def safe_format_date(date_obj, format_str: str = "%Y-%m-%d") -> str:
    """Format date dengan aman."""
    if date_obj is None:
        return ""

    try:
        if hasattr(date_obj, "strftime"):
            return date_obj.strftime(format_str)
        else:
            # Handle SQLAlchemy Date
            return str(date_obj)
    except (AttributeError, TypeError):
        return str(date_obj) if date_obj else ""


def safe_get_day(date_obj) -> int:
    """Get day dari date dengan aman."""
    if date_obj is None:
        return 1

    try:
        if hasattr(date_obj, "day"):
            return date_obj.day
        else:
            # Handle SQLAlchemy Date - convert dulu
            dt = safe_to_datetime(date_obj)
            return dt.day
    except (AttributeError, TypeError):
        return 1


def safe_relativedelta_operation(date_obj, delta_months: int):
    """Safe operation untuk relativedelta dengan date/datetime."""
    dt = safe_to_datetime(date_obj)
    return dt + relativedelta(months=delta_months)


router = APIRouter(
    prefix="/invoices",
    tags=["Invoices"],
    responses={404: {"description": "Not found"}},
)


# GET /invoices - Ambil semua data invoice
# Buat nampilin list invoice dengan fitur filter dan pencarian lengkap
# Query parameters:
# - search: cari berdasarkan nomor invoice, nama pelanggan, atau ID pelanggan
# - status_invoice: filter berdasarkan status (Belum Dibayar, Lunas, Kadaluarsa)
# - start_date: filter tanggal jatuh tempo mulai dari
# - end_date: filter tanggal jatuh tempo sampai dengan
# - show_active_only: kalo true, cuma tampilkan invoice yang punya payment_link aktif (90 hari terakhir)
# - skip: offset pagination (default: 0)
# - limit: jumlah data per halaman (default: semua)
# Response: list invoice dengan relasi lengkap (pelanggan, langganan, paket, data teknis)
# Performance: eager loading biar ga N+1 query
@router.get("/", response_model=List[InvoiceSchema])
async def get_all_invoices(
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user),  # PROTECTED
    search: Optional[str] = None,
    status_invoice: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    show_active_only: Optional[bool] = False,  
    exclude_expired: Optional[bool] = False, 
    skip: int = 0, 
    limit: Optional[int] = None,
):
    """Mengambil semua data invoice dengan filter."""
    # OPTIMIZED: Query dengan comprehensive eager loading untuk mencegah semua N+1 problems
    query = (
        select(InvoiceModel).join(InvoiceModel.pelanggan)
        # PERBAIKAN: Eager load semua relasi yang sering diakses bersama invoice
        # untuk mencegah N+1 queries saat data di-serialize ke response
        .options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.data_teknis),
                joinedload(PelangganModel.langganan).joinedload(LanggananModel.paket_layanan),
            )
        )
    )

    if search:
        search_term = f"%{search}%"
        # Cek apakah search term adalah angka (untuk pelanggan_id)
        is_numeric_search = search.strip().isdigit()

        query = query.where(
            or_(
                InvoiceModel.invoice_number.ilike(search_term),
                PelangganModel.nama.ilike(search_term),
                InvoiceModel.id_pelanggan.ilike(search_term),
                # Tambahkan filter langsung untuk pelanggan_id (BigInteger foreign key)
                # Ini memperbaiki bug di mana riwayat pembayaran di view langganan tidak muncul
                InvoiceModel.pelanggan_id == int(search.strip()) if is_numeric_search else False,
            )
        )

    if status_invoice:
        query = query.where(InvoiceModel.status_invoice == status_invoice)

    if start_date:
        query = query.where(InvoiceModel.tgl_jatuh_tempo >= start_date)
    if end_date:
        query = query.where(InvoiceModel.tgl_jatuh_tempo <= end_date)

    # Exclude invoice yang sudah kadaluarsa (lewat 5 hari grace period)
    if exclude_expired:
        from datetime import date
        overdue_threshold_date = date.today() - timedelta(days=5)
        query = query.where(
            or_(
                # Invoice Lunas selalu boleh tampil
                InvoiceModel.status_invoice == 'Lunas',
                # Invoice Belum Dibayar boleh tampil jika masih dalam grace period
                and_(
                    InvoiceModel.status_invoice == 'Belum Dibayar',
                    InvoiceModel.tgl_jatuh_tempo >= overdue_threshold_date
                )
            )
        )

    # Filter untuk hanya menampilkan invoice dengan link pembayaran aktif
    if show_active_only:
        # Hanya tampilkan invoice yang:
        # 1. Memiliki payment_link (baik yang sudah lunas maupun belum)
        # 2. Link pembayaran masih relevan (tidak terlalu tua)
        from datetime import date, timedelta

        cutoff_date = date.today() - timedelta(days=90)  # Invoice 90 hari terakhir

        query = query.where(
            and_(
                InvoiceModel.payment_link.isnot(None), InvoiceModel.payment_link != "", InvoiceModel.tgl_invoice >= cutoff_date
            )
        ).order_by(InvoiceModel.tgl_invoice.desc())

    # Optimized: Menambahkan order by untuk konsistent pagination
    if show_active_only:
        query = query.order_by(InvoiceModel.tgl_invoice.desc())
    else:
        query = query.order_by(InvoiceModel.created_at.desc())

    # Logic penentuan apakah perlu cek arsip
    # Kita cek arsip jika:
    # 1. Ada pencarian (search text)
    # 2. Filter status adalah 'Lunas', 'Kadaluarsa', atau 'Expired' (karena data ini biasanya diarsipkan)
    # 3. Rentang tanggal diberikan (karena mungkin data lama)
    check_archive = False
    if search:
        check_archive = True
    elif status_invoice in ['Lunas', 'Kadaluarsa', 'Expired']:
        check_archive = True
    elif start_date or end_date:
        check_archive = True

    # Jika perlu cek arsip, kita gunakan logic gabungan
    if check_archive:
        # Eksekusi query utama dulu (tanpa pagination limit/offset di sini)
        result_main = await db.execute(query)
        invoices_main = result_main.scalars().unique().all()

        # Query untuk arsip
        query_archive = (
            select(InvoiceArchiveModel).join(InvoiceArchiveModel.pelanggan)
            .options(
                joinedload(InvoiceArchiveModel.pelanggan)
            )
        )
        
        # Apply filters to archive query
        if search:
            search_term = f"%{search}%"
            query_archive = query_archive.where(
                or_(
                    InvoiceArchiveModel.invoice_number.ilike(search_term),
                    PelangganModel.nama.ilike(search_term),
                    InvoiceArchiveModel.id_pelanggan.ilike(search_term),
                )
            )
        
        if status_invoice:
            query_archive = query_archive.where(InvoiceArchiveModel.status_invoice == status_invoice)
            
        if start_date:
            query_archive = query_archive.where(InvoiceArchiveModel.tgl_jatuh_tempo >= start_date)
        if end_date:
            query_archive = query_archive.where(InvoiceArchiveModel.tgl_jatuh_tempo <= end_date)

        # Execute archive query
        result_archive = await db.execute(query_archive)
        invoices_archive = result_archive.scalars().unique().all()
        
        # Gabungkan hasil
        all_invoices = list(invoices_main) + list(invoices_archive)
        
        # Sortir gabungan (default by created_at desc, atau tgl_invoice desc)
        # Handle created_at is None in archive if any
        # Menggunakan tgl_invoice sebagai primary sort key karena lebih reliable untuk invoice lama
        all_invoices.sort(key=lambda x: x.tgl_invoice or x.created_at or datetime.min, reverse=True)
        
        # Manual Pagination slice
        start_idx = skip
        end_idx = skip + (limit if limit else len(all_invoices))
        return all_invoices[start_idx:end_idx]

    # Terapkan paginasi setelah semua filter (jika tidak cek arsip)
    query = query.order_by(InvoiceModel.created_at.desc()) # Ensure default sort
    query = query.offset(skip).limit(limit)
    # ---------------------------

    result = await db.execute(query)
    # FIX: Tambahkan .unique() untuk collection eager loading
    return result.scalars().unique().all()


def parse_xendit_datetime(iso_datetime_str: str) -> datetime:
    """Fungsi untuk mengkonversi format datetime ISO 8601 dari Xendit."""
    try:
        if not iso_datetime_str:
            pass  # This line was misplaced and caused an IndentationError. It's removed as it didn't have a clear purpose here.
        if iso_datetime_str.endswith("Z"):
            iso_datetime_str = iso_datetime_str[:-1] + "+00:00"
        return datetime.fromisoformat(iso_datetime_str)
    except (ValueError, TypeError):
        return datetime.now(timezone.utc)


async def _process_successful_payment(db: AsyncSession, invoice: InvoiceModel, payload: dict | None = None):
    """Fungsi terpusat untuk menangani logika setelah invoice lunas."""

    pelanggan = invoice.pelanggan
    if not pelanggan or not pelanggan.langganan:
        logger.error(f"Pelanggan atau langganan tidak ditemukan untuk invoice {invoice.invoice_number}")
        return

    langganan = pelanggan.langganan[0]

    # Cek apakah langganan sebelumnya berstatus 'Suspended'
    is_suspended_or_inactive = langganan.status == "Suspended" or not langganan.status

    # Update status invoice (tetap)
    invoice.status_invoice = "Lunas"
    if payload:
        invoice.paid_amount = float(payload.get("paid_amount", invoice.total_harga or 0))
        paid_at_str = payload.get("paid_at")
        invoice.paid_at = parse_xendit_datetime(paid_at_str) if paid_at_str else datetime.now(timezone.utc)
    else:
        invoice.paid_amount = invoice.total_harga
        invoice.paid_at = datetime.now(timezone.utc)

    db.add(invoice)

    next_due_date = None

    if langganan.metode_pembayaran == "Prorate":
        paket = langganan.paket_layanan
        brand = pelanggan.harga_layanan
        langganan.metode_pembayaran = "Otomatis"
        current_due_date = invoice.tgl_jatuh_tempo

        if not paket or not brand:
            logger.error(f"Data paket/brand tidak lengkap untuk langganan ID {langganan.id}")
            # Set fallback jika data tidak ada, agar tidak crash
            current_due_datetime = safe_to_datetime(current_due_date)
            next_due_datetime = current_due_datetime + relativedelta(months=1)
            next_due_date = next_due_datetime.date().replace(day=1)
        else:
            # Hitung harga normal penuh sebagai pembanding
            harga_paket = float(paket.harga)
            pajak_persen = float(brand.pajak)
            harga_normal_full = harga_paket * (1 + (pajak_persen / 100))

            # Logika Pembeda: Apakah ini tagihan prorate biasa atau gabungan?
            if float(invoice.total_harga or 0) > (harga_normal_full + 1):
                # Skenario 1: INI ADALAH TAGIHAN GABUNGAN
                # Convert Date to datetime untuk relativedelta - handle SQLAlchemy Date
                current_due_datetime = safe_to_datetime(current_due_date)
                
                if current_due_datetime.day == 1:
                    # Jika jatuh tempo tgl 1 (misal 1 Maret), berarti cover Feb & Maret.
                    # Next due: 1 April. (Current + 1 bulan)
                    next_due_datetime = current_due_datetime + relativedelta(months=1)
                else:
                    # Jika jatuh tempo tgl 28 (misal 28 Feb), cover Feb & Maret.
                    # Next due: 1 April. (Current + 2 bulan di-force tanggal 1)
                    next_due_datetime = current_due_datetime + relativedelta(months=2)
                
                next_due_date = next_due_datetime.date().replace(day=1)
                logger.info(f"Tagihan gabungan terdeteksi. Jatuh tempo berikutnya diatur ke {next_due_date}")
            else:
                # Skenario 2: INI ADALAH TAGIHAN PRORATE BIASA
                # Convert Date to datetime untuk relativedelta - handle SQLAlchemy Date
                current_due_datetime = safe_to_datetime(current_due_date)
                
                if current_due_datetime.day == 1:
                    # Jika jatuh tempo tgl 1 (misal 1 Maret), ini adalah pembayaran untuk Feb.
                    # Siklus Maret dimulai 1 Maret. Jadi Next Due tetap 1 Maret (agar tagihan Maret ter-generate/aktif).
                    next_due_date = current_due_datetime.date()
                else:
                    # Jika jatuh tempo tgl 28 (misal 28 Feb), pembayaran Feb.
                    # Next due: 1 Maret (Current + 1 bulan di-force tanggal 1).
                    next_due_datetime = current_due_datetime + relativedelta(months=1)
                    next_due_date = next_due_datetime.date().replace(day=1)
                    
                logger.info(f"Tagihan prorate biasa terdeteksi. Jatuh tempo berikutnya diatur ke {next_due_date}")

            # Reset harga langganan ke harga normal untuk bulan-bulan berikutnya
            langganan.harga_awal = round(harga_normal_full, 0)

    else:  # Skenario 3: Jika sudah Otomatis (PEMBAYARAN BULANAN NORMAL)
        # FIX LOGIC: Gunakan invoice.tgl_jatuh_tempo sebagai basis perhitungan, BUKAN langganan.tgl_jatuh_tempo.
        # Menggunakan langganan.tgl_jatuh_tempo (mutable) berisiko jika callback diproses 2x (misal user bayar double klik atau retry),
        # yang menyebabkan tanggal loncat 2 bulan (Des -> Jan -> Feb).
        # Dengan pakai invoice.tgl_jatuh_tempo (fixed), hasilnya selalu konsisten (Des + 1 bulan = Jan).
        current_due_date = invoice.tgl_jatuh_tempo
        
        # Fallback safety jika data lama tidak punya tgl_jatuh_tempo
        if not current_due_date:
            current_due_date = date.today()
            logger.warning(f"Invoice {invoice.invoice_number} tidak punya tgl_jatuh_tempo, pakai today: {current_due_date}")

        # Convert Date to datetime untuk relativedelta - handle SQLAlchemy Date
        current_due_datetime = safe_to_datetime(current_due_date)
        next_due_datetime = current_due_datetime + relativedelta(months=1)
        next_due_date = next_due_datetime.date().replace(day=1)

    # Update langganan (tidak berubah)
    langganan.status = "Aktif"
    langganan.tgl_jatuh_tempo = next_due_date
    langganan.tgl_invoice_terakhir = date.today()
    db.add(langganan)

    # HANYA trigger update Mikrotik jika status sebelumnya adalah 'Suspended'
    if is_suspended_or_inactive:
        data_teknis = pelanggan.data_teknis
        if not data_teknis:
            logger.error(f"Data Teknis tidak ditemukan untuk pelanggan ID {pelanggan.id}. Mikrotik update dilewati.")
        else:
            try:
                # Panggil fungsi dengan SEMUA argumen yang dibutuhkan
                await mikrotik_service.trigger_mikrotik_update(
                    db,
                    langganan,
                    data_teknis,
                    data_teknis.id_pelanggan,  # old_id_pelanggan diisi dengan id saat ini
                )
                logger.info(f"Berhasil trigger re-aktivasi Mikrotik untuk langganan ID {langganan.id}")

                # Jika berhasil, set flag pending sync (jika ada) kembali ke False
                if data_teknis.mikrotik_sync_pending:
                    data_teknis.mikrotik_sync_pending = False
                    db.add(data_teknis)

            except Exception as e:
                # Jika GAGAL, catat error DAN set flag retry menjadi True
                logger.error(
                    f"Gagal trigger re-aktivasi Mikrotik untuk langganan ID {langganan.id}: {e}. Menandai untuk dicoba lagi."
                )
                data_teknis.mikrotik_sync_pending = True
                db.add(data_teknis)
    else:
        logger.info(f"Langganan ID {langganan.id} sudah Aktif. Mikrotik update dilewati.")

    # Notif ke frontend
    try:
        target_roles = ["Admin", "NOC", "Finance"]
        query = select(UserModel.id).join(RoleModel).where(func.lower(RoleModel.name).in_([r.lower() for r in target_roles]))
        result = await db.execute(query)
        target_user_ids = result.scalars().all()

        if target_user_ids:
            # Pastikan pelanggan sudah di-load dengan benar
            pelanggan_nama = pelanggan.nama if pelanggan else "N/A"
            notification_payload = {
                "type": "new_payment",
                "message": f"Pembayaran untuk invoice {invoice.invoice_number} dari {pelanggan_nama} telah diterima.",
                "timestamp": datetime.now().isoformat(),
                "data": {
                    "invoice_id": invoice.id,
                    "invoice_number": invoice.invoice_number,
                    "pelanggan_nama": pelanggan_nama,
                    "amount": (float(invoice.total_harga) if invoice.total_harga else 0.0),
                    "payment_method": invoice.metode_pembayaran or "Unknown",
                    "timestamp": datetime.now().isoformat(),
                },
            }
            # Menambahkan log ini untuk memastikan user ID ditemukan
            logger.info(f"Mencoba mengirim notifikasi pembayaran ke user IDs: {target_user_ids}")
            # Convert ke list untuk broadcast_to_roles
            user_ids_list = list(target_user_ids)
            await manager.broadcast_to_roles(notification_payload, user_ids_list)
            logger.info(f"Notifikasi pembayaran berhasil dikirim untuk invoice {invoice.invoice_number}")
        else:
            logger.warning(f"Tidak ada user dengan role Admin/CS yang ditemukan untuk dikirimi notifikasi.")

    except Exception as e:
        # 🛡️ Graceful degradation: Payment processed but notification failed
        logger.error(f"⚠️ Payment successful but notification failed for invoice {invoice.invoice_number}: {e}", exc_info=True)
        # Continue processing - payment is still valid even if notification fails

    logger.info(f"Payment processed successfully for invoice {invoice.invoice_number}")


# @router.post("/xendit-callback", status_code=status.HTTP_200_OK)
# async def handle_xendit_callback(
#     request: Request,
#     x_callback_token: Optional[str] = Header(None),
#     db: AsyncSession = Depends(get_db),
# ):
#     payload = await request.json()
#     logger.info(f"Xendit callback received. Payload: {json.dumps(payload, indent=2)}")

#     external_id = payload.get("external_id")
#     if not external_id:
#         raise HTTPException(
#             status_code=400, detail="External ID tidak ditemukan di payload"
#         )

#     try:
#         brand_prefix = external_id.split("/")[0]
#     except IndexError:
#         raise HTTPException(status_code=400, detail="Format external_id tidak valid")

#     correct_token = None
#     if brand_prefix.lower() in ["jakinet", "nagrak"]:
#         correct_token = settings.XENDIT_CALLBACK_TOKENS.get("ARTACOM")
#         logger.info("Validating with ARTACOM callback token.")
#     elif brand_prefix.lower() == "jelantik":
#         correct_token = settings.XENDIT_CALLBACK_TOKENS.get("JELANTIK")
#         logger.info("Validating with JELANTIK callback token.")

#     if not correct_token or x_callback_token != correct_token:
#         logger.warning(f"Invalid callback token received for brand '{brand_prefix}'.")
#         raise HTTPException(
#             status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid callback token"
#         )


# POST /invoices/xendit-callback - Callback dari Xendit payment gateway
# Endpoint buat terima callback dari Xendit setelah customer bayar invoice
# Request headers:
# - x_callback_token: token untuk validasi callback (sesuai brand)
# Request body: payment data dari Xendit (JSON)
# Response: success message
# Security:
# - Validasi callback token berdasarkan brand prefix
# - Cek duplikasi callback biar idempotent
# - Log semua callback untuk audit trail
# Fitur:
# - Update status invoice jadi "Lunas" kalo payment sukses
# - Update status jadi "Kadaluarsa" kalo payment expired
# - Trigger re-aktivasi Mikrotik kalo pelanggan sebelumnya suspended
# - Kirim notifikasi ke Admin/NOC/Finance
# Error handling: graceful degradation, payment tetep valid walau notifikasi gagal
@router.post("/xendit-callback", status_code=status.HTTP_200_OK)
async def handle_xendit_callback(
    request: Request,
    x_callback_token: Optional[str] = Header(None),
    db: AsyncSession = Depends(get_db),
):
    # Get raw body for logging but filter sensitive data
    raw_body = await request.body()

    # Log the callback with filtered data
    try:
        if raw_body:
            body_str = raw_body.decode("utf-8")
            filtered_body = sanitize_log_data(body_str)
            logger.info(f"Xendit callback received. Filtered Payload: {filtered_body}")
        else:
            logger.info("Xendit callback received. Payload: Empty body")
    except Exception as e:
        logger.info(f"Xendit callback received. Payload: ***REDACTED*** (Error processing body: {str(e)})")

    payload = await request.json()
    # Log the JSON payload with sensitive data filtered
    filtered_payload = sanitize_log_data(payload)
    logger.info(f"Xendit callback received. Filtered JSON Payload: {json.dumps(filtered_payload, indent=2)}")

    # Extract IDs from payload
    xendit_id = payload.get("id")  # Xendit internal ID
    external_id = payload.get("external_id")
    xendit_status = payload.get("status")

    # Extract idempotency key if provided in headers
    idempotency_key = request.headers.get("x-idempotency-key", request.headers.get("idempotency-key"))

    if not external_id:
        raise HTTPException(status_code=400, detail="External ID tidak ditemukan di payload")

    # Check for duplicate callback
    is_duplicate = await check_duplicate_callback(db, xendit_id or external_id, external_id, idempotency_key or "")
    if is_duplicate:
        logger.info(f"Duplicated callback received and ignored: xendit_id={xendit_id}, external_id={external_id}")
        return {"message": "Callback already processed"}

    brand_prefix = None
    try:
        if "/" in external_id:
            brand_prefix = external_id.split("/")[0]
        # Untuk webhook test, tidak ada prefix brand. Kita validasi token saja
    except IndexError:
        raise HTTPException(status_code=400, detail="Format external_id tidak valid")

    # VALIDASI TOKEN SECARA LANGSUNG
    correct_token = None  # FIX: Inisialisasi untuk mencegah NameError jika tidak ada token yang cocok
    # 1. Coba validasi dengan token ARTACOMINDO (Jakinet, Nagrak)
    artacom_token = settings.XENDIT_CALLBACK_TOKENS.get("ARTACOMINDO")
    if artacom_token and x_callback_token == artacom_token:
        logger.info("Validating with ARTACOMINDO callback token.")
        # Cek apakah brand_prefix (jika ada) sesuai dengan token ini
        if brand_prefix and brand_prefix.lower() not in [
            "jakinet",
            "nagrak",
            "artacom",
        ]:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid brand for this token",
            )
        correct_token = artacom_token

    # 2. Coba validasi dengan token JELANTIK
    jelantik_token = settings.XENDIT_CALLBACK_TOKENS.get("JELANTIK")
    if jelantik_token and x_callback_token == jelantik_token:
        logger.info("Validating with JELANTIK callback token.")
        # Cek apakah brand_prefix (jika ada) sesuai dengan token ini
        if brand_prefix and brand_prefix.lower() != "jelantik":
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid brand for this token",
            )
        correct_token = jelantik_token

    # Jika tidak ada token yang cocok, kembalikan 401
    if not correct_token:
        logger.warning("Invalid callback token received.")
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid callback token")

    # Log callback processing to prevent duplicates
    logged = await log_callback_processing(
        db,
        xendit_id or external_id,  # Use xendit_id if available, otherwise use external_id
        external_id,
        xendit_status,
        payload,
        idempotency_key or "",
    )

    # If logging failed, it means another process already handled this callback
    if not logged:
        logger.info(
            f"Callback already processed by another concurrent request: xendit_id={xendit_id}, external_id={external_id}"
        )
        return {"message": "Callback already processed"}

    # Proceed with normal processing
    # Buat filter conditions untuk query invoice
    # Cari berdasarkan xendit_external_id terlebih dahulu, kemudian fallback ke invoice_number
    filter_conditions = [InvoiceModel.xendit_external_id == external_id, InvoiceModel.invoice_number == external_id]

    # Logging tambahan untuk debugging
    logger.info(f"Searching for invoice with external_id: {external_id}")

    # Optimasi query dengan menggunakan joinedload untuk relasi yang sering digunakan bersama
    # Ini akan menghindari N+1 query problem
    stmt = (
        select(InvoiceModel)
        .join(InvoiceModel.pelanggan)
        .options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.langganan).joinedload(LanggananModel.paket_layanan),
                joinedload(PelangganModel.data_teknis),
            )
        )
        .where(or_(*filter_conditions))
    )
    invoice = (await db.execute(stmt)).unique().scalar_one_or_none()

    # Logging tambahan untuk debugging
    if not invoice:
        # Coba cari dengan LIKE untuk melihat apakah ada invoice yang mirip
        search_stmt = select(InvoiceModel).where(
            or_(InvoiceModel.invoice_number.like(f"%{external_id}%"), InvoiceModel.xendit_external_id.like(f"%{external_id}%"))
        )
        similar_invoices = (await db.execute(search_stmt)).scalars().all()
        if similar_invoices:
            logger.info(f"Found similar invoices: {[inv.invoice_number for inv in similar_invoices]}")
        else:
            logger.info("No similar invoices found")

    if not invoice:
        logger.warning(f"Invoice with external_id {external_id} not found, but callback is valid.")
        return {"message": "Callback valid, invoice not found."}

    if invoice.status_invoice == "Lunas":
        # Still check if this callback was already logged for idempotency tracking
        # This can happen if the invoice status was updated but the callback logging failed
        callback_exists = await check_duplicate_callback(db, xendit_id or external_id, external_id, idempotency_key or "")
        if not callback_exists:
            # If no callback record exists, create one for tracking purposes
            await log_callback_processing(
                db,
                xendit_id or external_id,  # Use xendit_id if available, otherwise use external_id
                external_id,
                xendit_status,
                payload,
                idempotency_key or "",
            )

        logger.info(f"Invoice {invoice.invoice_number} already has status Lunas, callback ignored.")
        return {"message": "Invoice already processed"}

    # SAFETY NET: Simpan status original untuk potensi Mikrotik rollback
    _original_langganan_status = None
    _rollback_langganan = None
    _rollback_data_teknis = None
    if invoice.pelanggan and invoice.pelanggan.langganan:
        _rollback_langganan = invoice.pelanggan.langganan[0]
        _original_langganan_status = _rollback_langganan.status
        _rollback_data_teknis = getattr(invoice.pelanggan, 'data_teknis', None)

    try:
        if xendit_status == "PAID":
            await _process_successful_payment(db, invoice, payload)

        elif xendit_status == "EXPIRED":
            invoice.status_invoice = "Kadaluarsa"
            db.add(invoice)

        await db.commit()

        # Clear sidebar cache untuk update badge count
        try:
            from .dashboard import clear_sidebar_cache
            clear_sidebar_cache()
        except ImportError:
            pass  # Fallback jika import gagal
    except Exception as e:
        await db.rollback()
        logger.error(f"Error processing Xendit callback for external_id {external_id}: {str(e)}")

        # SAFETY NET: Rollback Mikrotik jika unsuspend berhasil tapi commit gagal
        if (xendit_status == "PAID" and _original_langganan_status == "Suspended"
                and _rollback_langganan and _rollback_data_teknis):
            try:
                logger.warning(f"🔄 Attempting Mikrotik rollback for langganan ID {_rollback_langganan.id}...")
                _rollback_langganan.status = "Suspended"
                await mikrotik_service.trigger_mikrotik_update(
                    db, _rollback_langganan, _rollback_data_teknis, _rollback_data_teknis.id_pelanggan
                )
                logger.info(f"✅ Mikrotik rollback successful for langganan ID {_rollback_langganan.id}")
            except Exception as rollback_error:
                logger.error(f"❌ CRITICAL: Mikrotik rollback FAILED: {rollback_error}")
                _rollback_data_teknis.mikrotik_sync_pending = True
                try:
                    db.add(_rollback_data_teknis)
                    await db.commit()
                except Exception:
                    logger.error("⚠️ Could not mark mikrotik_sync_pending. Manual intervention required.")

        raise HTTPException(status_code=500, detail="Internal server error while processing callback.")

    return {"message": "Callback processed successfully"}


# POST /invoices/create_reinvoice/{expired_invoice_id} - Buat reinvoice dari invoice expired
# Buat reinvoice dari invoice yang sudah expired
# Path parameters:
# - expired_invoice_id: ID invoice yang mau direinvoice
# Response: data invoice baru + invoice lama yang diupdate
# Permission: butuh login (authenticated user)
# Fitur:
# - Copy data dari invoice expired
# - Update tanggal jatuh tempo ke bulan depan
# - Set flag is_reinvoice=True
# - Link ke invoice asli via original_invoice_id
# - Update status invoice lama jadi Expired
# Validation:
# - Cek invoice expired ada
# - Cek pelanggan punya langganan aktif
# Error handling: 404 kalo invoice/langganan nggak ada, 500 kalo gagal create
@router.post("/create_reinvoice/{expired_invoice_id}")
async def create_reinvoice(
    expired_invoice_id: int,
    current_user: UserModel = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Buat reinvoice dari invoice yang sudah expired.
    Otomatis menyalin data dan menambahkan flag reinvoice.
    """
    try:
        # Query invoice yang expired
        stmt = select(InvoiceModel).where(InvoiceModel.id == expired_invoice_id).options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.langganan),
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.data_teknis),
            )
        )
        result = await db.execute(stmt)
        expired_invoice = result.unique().scalar_one_or_none()

        if not expired_invoice:
            raise HTTPException(status_code=404, detail="Invoice tidak ditemukan")

        # Cari langganan untuk pelanggan ini (TIDAK PEDULI status - bisa Aktif/Suspended/Suspend)
        # Reinvoice harus bisa dibuat untuk langganan dengan status apapun
        langganan = None
        if expired_invoice.pelanggan and expired_invoice.pelanggan.langganan:
            # Ambil langganan pertama yang ada (biasanya pelanggan hanya punya 1 langganan aktif)
            # Prioritas: Aktif > Suspended > Suspend > Berhenti
            status_priority = {"Aktif": 1, "Suspended": 2, "Suspend": 3, "Berhenti": 4}
            sorted_langganan = sorted(
                expired_invoice.pelanggan.langganan,
                key=lambda x: status_priority.get(x.status, 99)
            )
            langganan = sorted_langganan[0] if sorted_langganan else None

        if not langganan:
            raise HTTPException(status_code=404, detail="Tidak ada langganan untuk pelanggan ini")

        # Update tanggal jatuh tempo ke bulan depan
        from datetime import date
        from dateutil.relativedelta import relativedelta

        current_due = date.fromisoformat(str(expired_invoice.tgl_jatuh_tempo))
        new_due = current_due + relativedelta(months=1)
        langganan.tgl_jatuh_tempo = new_due

        # Buat data invoice generate
        # generate_manual_invoice akan otomatis menghitung harga berdasarkan metode_pembayaran langganan
        # (Otomatis = harga penuh, Prorate = harga proporsional)
        from ..schemas.invoice import InvoiceGenerate
        invoice_data = InvoiceGenerate(
            langganan_id=langganan.id,
            is_reinvoice=True,
            original_invoice_id=expired_invoice_id,
            reinvoice_reason="expired"
        )

        # Generate invoice baru - panggil fungsi yang sudah ada
        new_invoice = await generate_manual_invoice(invoice_data, db)

        # Update status invoice lama menjadi Expired jika belum
        if expired_invoice.status_invoice != "Expired":
            expired_invoice.status_invoice = "Expired"

        await db.commit()

        # Convert ke dictionary untuk response
        return {
            "message": "Reinvoice berhasil dibuat",
            "new_invoice": {
                "id": new_invoice.id,
                "invoice_number": new_invoice.invoice_number,
                "status_invoice": new_invoice.status_invoice,
                "total_harga": new_invoice.total_harga,
                "tgl_jatuh_tempo": str(new_invoice.tgl_jatuh_tempo),
                "is_reinvoice": new_invoice.is_reinvoice,
                "original_invoice_id": new_invoice.original_invoice_id
            },
            "expired_invoice": {
                "id": expired_invoice.id,
                "invoice_number": expired_invoice.invoice_number,
                "status_invoice": expired_invoice.status_invoice
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating reinvoice: {e}")
        await db.rollback()
        raise HTTPException(status_code=500, detail="Gagal membuat reinvoice")


# POST /invoices/generate - Generate invoice manual
# Buat bikin invoice baru secara manual berdasarkan langganan yang udah ada
# Request body: langganan_id
# Response: data invoice yang baru dibuat dengan payment link dari Xendit
# Permission: butuh permission "create_invoices"
# Fitur:
# - Hitung harga otomatis berdasarkan paket dan pajak
# - Generate nomor invoice otomatis
# - Create payment link di Xendit
# - Dynamic description (prorate/bulan penuh)
# - Format nomor telepon buat Xendit (0xx -> 62xx)
# Validation:
# - Cek langganan harus ada
# - Cek pelanggan harus ada
# - Cek status langganan (boleh: Aktif, Suspended)
# - Cek duplicate invoice di bulan yang sama
# Error handling: rollback transaction kalo ada error
@router.post(
    "/generate",
    response_model=InvoiceSchema,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(has_permission("create_invoices"))],
)
async def generate_manual_invoice(invoice_data: InvoiceGenerate, db: AsyncSession = Depends(get_db)):
    """Membuat satu invoice secara manual berdasarkan langganan_id."""
    # PERBAIKAN: Query dengan eager loading yang lebih robust untuk mencegah N+1 problems
    stmt = (
        select(LanggananModel)
        .where(LanggananModel.id == invoice_data.langganan_id)
        .options(
            joinedload(LanggananModel.pelanggan).options(
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.data_teknis),
            ),
            joinedload(LanggananModel.paket_layanan),
        )
    )
    result = await db.execute(stmt)
    langganan = result.unique().scalar_one_or_none()

    if not langganan:
        raise HTTPException(status_code=404, detail=f"Langganan dengan ID {invoice_data.langganan_id} tidak ditemukan")

    # VALIDASI: Periksa apakah pelanggan dari langganan benar-benar ada
    if not langganan.pelanggan:
        logger.error(
            f"Langganan ID {langganan.id} memiliki pelanggan_id {langganan.pelanggan_id} tapi data pelanggan tidak ditemukan di database"
        )
        raise HTTPException(
            status_code=404,
            detail=f"Pelanggan dengan ID {langganan.pelanggan_id} tidak ditemukan untuk langganan ID {langganan.id}. Data mungkin tidak konsisten.",
        )

    if langganan.status == "Berhenti":
        raise HTTPException(
            status_code=400,
            detail=f"Gagal membuat invoice. Status langganan untuk pelanggan '{langganan.pelanggan.nama}' adalah 'Berhenti'.",
        )

    pelanggan = langganan.pelanggan
    paket = langganan.paket_layanan
    if not pelanggan or not paket or not pelanggan.harga_layanan or not pelanggan.data_teknis:
        missing_data = []
        if not pelanggan:
            missing_data.append("pelanggan")
        if not paket:
            missing_data.append("paket_layanan")
        if not pelanggan.harga_layanan:
            missing_data.append("harga_layanan/brand")
        if not pelanggan.data_teknis:
            missing_data.append("data_teknis")

        raise HTTPException(
            status_code=400,
            detail=f"Data pendukung tidak lengkap untuk langganan ID {langganan.id}: {', '.join(missing_data)}",
        )

    brand = pelanggan.harga_layanan
    data_teknis = pelanggan.data_teknis

    if not paket.harga or not brand.pajak:
        raise HTTPException(status_code=400, detail="Harga paket atau pajak tidak valid")

    # Target jatuh tempo untuk database
    # Jika Prorate dan jatuh temponya tanggal 1, kita simpan sebagai tanggal terakhir bulan sebelumnya
    # agar sesuai dengan keinginan Finance (visual 28 Feb bukan 1 Mar)
    actual_due_date = langganan.tgl_jatuh_tempo
    if langganan.metode_pembayaran == "Prorate" and actual_due_date.day == 1:
        from datetime import timedelta
        actual_due_date = actual_due_date - timedelta(days=1)

    # Check existing invoice - tapi allow jika status expired/kadaluarsa dan ini adalah reinvoice
    if not invoice_data.is_reinvoice:
        existing_invoice_stmt = select(InvoiceModel.id).where(
            InvoiceModel.pelanggan_id == langganan.pelanggan_id,
            InvoiceModel.tgl_jatuh_tempo == actual_due_date,
        )
        existing = (await db.execute(existing_invoice_stmt)).scalar_one_or_none()
        if existing:
            raise HTTPException(status_code=409, detail="Invoice untuk periode ini sudah ada.")

    # Handle SQLAlchemy Date untuk formatting
    jatuh_tempo_date = langganan.tgl_jatuh_tempo
    jatuh_tempo_str = safe_format_date(jatuh_tempo_date, "%d/%m/%Y")
    jatuh_tempo_yyyymm = safe_format_date(jatuh_tempo_date, "%Y%m") or "202501"  # Default fallback

    # --- MODIFICATION FOR INVOICE NUMBER ---
    # 1. Sanitize customer name and address
    import re
    nama_pelanggan_singkat = re.sub(r'[^a-zA-Z0-9]', '', pelanggan.nama).upper()
    alamat_singkat = re.sub(r'[^a-zA-Z0-9]', '', pelanggan.alamat or '').upper()[:10]  # Take only first 10 chars
    brand_singkat = re.sub(r'[^a-zA-Z0-9]', '', brand.brand or '').upper()
    # Convert SQLAlchemy Date to Python date for datetime.combine
    from datetime import date, datetime, timedelta
    jatuh_tempo_python_date = date.fromisoformat(str(langganan.tgl_jatuh_tempo))
    
    # Logic penentuan nama bulan untuk Nomor Invoice
    # Jika Prorate dan jatuh tempo tgl 1, gunakan bulan sebelumnya untuk penamaan
    naming_date = jatuh_tempo_python_date
    if langganan.metode_pembayaran == "Prorate" and naming_date.day == 1:
        naming_date = naming_date - timedelta(days=1)
        
    bulan_tahun = datetime.combine(naming_date, datetime.min.time()).strftime("%B-%Y").upper()

    # 2. Generate ID suffix - gunakan id_pelanggan jika ada, fallback ke invoice.id
    # Ini mencegah issue "external_id" berujung "None" di Xendit (contoh: .../PARAMA/None)
    if data_teknis.id_pelanggan:
        id_suffix = str(data_teknis.id_pelanggan)[-3:]  # 3 digit terakhir id_pelanggan
    else:
        # Fallback: gunakan temporary placeholder yang akan diganti setelah invoice dibuat
        id_suffix = "TMP"
    
    # 3. Generate new invoice number in format: BRAND/LAYANAN/NAMA_PELANGGAN/BULAN_TAHUN/ALAMAT_SINGKAT/ID_SUFFIX
    nomor_invoice = f"{brand_singkat}/ftth/{nama_pelanggan_singkat}/{bulan_tahun}/{alamat_singkat}/{id_suffix}"

    # 4. Check for duplicate invoice number and add timestamp if needed
    existing_invoice_number = (await db.execute(
        select(InvoiceModel.id).where(InvoiceModel.invoice_number == nomor_invoice)
    )).scalar_one_or_none()

    if existing_invoice_number:
        # Generate nomor unik dengan tambahan timestamp atau random
        import time
        timestamp = str(int(time.time()))[-6:]  # 6 digit terakhir timestamp
        nomor_invoice = f"{nomor_invoice}/{timestamp}"
    # --- END OF MODIFICATION ---

    # Ambil total harga langsung dari data langganan yang sudah dihitung (prorate + PPN).
    total_harga = float(langganan.harga_awal or 0)
    pajak_persen = float(brand.pajak or 0)

    # Karena Xendit butuh nilai pajak terpisah, kita hitung mundur dari total harga.
    # 1. Cari harga dasar sebelum pajak.
    harga_dasar = total_harga / (1 + (pajak_persen / 100))

    # 2. Hitung nilai pajak berdasarkan selisih total harga dan harga dasar.
    # Gunakan pembulatan untuk konsistensi.
    pajak = round(total_harga - harga_dasar)

    # Pastikan harga dasar untuk item di Xendit juga konsisten.
    harga_dasar = total_harga - pajak

    # Simpan harga sebelum diskon untuk reference
    total_harga_sebelum_diskon = total_harga

    # Cek diskon aktif untuk cluster pelanggan
    diskon_applied = None
    diskon_id = None
    diskon_persen = None
    diskon_amount = None

    # Gunakan alamat sebagai cluster untuk diskon
    cluster_to_check = pelanggan.alamat if pelanggan.alamat and pelanggan.alamat.strip() else None

    # Prorate users TIDAK mendapatkan diskon
    if langganan.metode_pembayaran == "Prorate":
        logger.info(f"⚠️ Invoice manual untuk pelanggan {pelanggan.nama} (ID: {pelanggan.id}) menggunakan Prorate - Diskon tidak diterapkan")
    elif cluster_to_check:
        # Cari diskon aktif untuk cluster ini
        tanggal_hari_ini = date.today()
        diskon_query = (
            select(DiskonModel)
            .where(DiskonModel.cluster == cluster_to_check)
            .where(DiskonModel.is_active == True)
            .where(
                (DiskonModel.tgl_mulai.is_(None)) | (DiskonModel.tgl_mulai <= tanggal_hari_ini)
            )
            .where(
                (DiskonModel.tgl_selesai.is_(None)) | (DiskonModel.tgl_selesai >= tanggal_hari_ini)
            )
            .order_by(DiskonModel.persentase_diskon.desc())
        )
        diskon_result = await db.execute(diskon_query)
        diskon_applied = diskon_result.scalar_one_or_none()

    if diskon_applied:
        diskon_id = diskon_applied.id
        diskon_persen = float(diskon_applied.persentase_diskon)
        diskon_amount = math.floor((total_harga_sebelum_diskon * diskon_persen / 100) + 0.5)
        total_harga = total_harga_sebelum_diskon - diskon_amount
        logger.info(f"💰 Diskon {diskon_persen}% (Rp {diskon_amount:,.0f}) diterapkan untuk invoice manual pelanggan {pelanggan.nama} - Cluster: {cluster_to_check}")

    # Tentukan tipe invoice
    invoice_type = "manual"
    if invoice_data.is_reinvoice:
        invoice_type = "reinvoice"

    new_invoice_data = {
            "invoice_number": nomor_invoice,
            "pelanggan_id": pelanggan.id,
            "id_pelanggan": data_teknis.id_pelanggan,
            "brand": brand.brand,
            "total_harga": total_harga,
            "no_telp": pelanggan.no_telp,
            "email": pelanggan.email,
            "tgl_invoice": date.today(),
            "tgl_jatuh_tempo": actual_due_date, # Menggunakan actual_due_date yang sudah disesuaikan (misal 28 Feb)
            "status_invoice": "Belum Dibayar",
            # Tipe invoice
            "invoice_type": invoice_type,
            # Menambahkan field reinvoice tracking
            "is_reinvoice": invoice_data.is_reinvoice if hasattr(invoice_data, 'is_reinvoice') else False,
            "original_invoice_id": invoice_data.original_invoice_id if hasattr(invoice_data, 'original_invoice_id') else None,
            "reinvoice_reason": invoice_data.reinvoice_reason if hasattr(invoice_data, 'reinvoice_reason') else None,
            # Diskon fields
            "diskon_id": diskon_id,
            "diskon_persen": diskon_persen,
            "diskon_amount": diskon_amount,
            "harga_sebelum_diskon": total_harga_sebelum_diskon if diskon_applied else None,
        }

    db_invoice = InvoiceModel(**new_invoice_data)
    db.add(db_invoice)
    await db.flush()
    
    # Jika menggunakan placeholder TMP (karena id_pelanggan None), ganti dengan invoice.id
    if id_suffix == "TMP":
        # Update invoice_number dengan invoice.id yang sebenarnya
        id_suffix_final = str(db_invoice.id)[-3:].zfill(3)  # 3 digit terakhir, pad dengan 0 jika perlu
        db_invoice.invoice_number = f"{brand_singkat}/ftth/{nama_pelanggan_singkat}/{bulan_tahun}/{alamat_singkat}/{id_suffix_final}"
        logger.warning(f"⚠️ id_pelanggan is None for {pelanggan.nama}, using invoice.id instead: {db_invoice.invoice_number}")


    try:
        deskripsi_xendit = ""
        # Handle SQLAlchemy Date untuk formatting
        due_date_obj = db_invoice.tgl_jatuh_tempo
        jatuh_tempo_str_lengkap = safe_format_date(due_date_obj, "%d/%m/%Y")

        if langganan.metode_pembayaran == "Prorate":

            # Hitung harga normal untuk perbandingan
            harga_normal_full = float(paket.harga) * (1 + (float(brand.pajak or 0) / 100))

            # Define invoice_date and due_date before they are referenced
            invoice_date_obj = db_invoice.tgl_invoice
            due_date_obj = db_invoice.tgl_jatuh_tempo

            # Convert ke Python date/datetime dengan aman
            invoice_date = safe_to_datetime(invoice_date_obj) if invoice_date_obj else datetime.now()
            due_date = safe_to_datetime(due_date_obj) if due_date_obj else datetime.now()

            # FIX: Untuk Prorate, periode harus dihitung berdasarkan due_date (tgl_jatuh_tempo)
            # karena due_date adalah akhir dari periode yang ditagihkan
            # Mulai periode = due_date - 1 bulan + 1 hari (awal bulan)
            # Akhir periode = due_date (tanggal jatuh tempo)
            due_date_only = due_date.date() if hasattr(due_date, 'date') else due_date

            # Cek apakah ini invoice pertama (pelanggan baru)
            # Invoice pertama: tgl_invoice sama dengan tgl mulai langganan (bukan tanggal 1)
            is_first_invoice = invoice_date.day != 1

            if is_first_invoice:
                # Invoice pertama: gunakan tgl_invoice sebagai mulai periode
                periode_start = invoice_date
                
                # Jika jatuh tempo tanggal 1, user ingin display-nya akhir bulan sebelumnya (misal 28 Feb)
                # agar tidak membingungkan pelanggan (seolah-olah 1 hari di bulan baru)
                if due_date.day == 1:
                    periode_end = due_date - timedelta(days=1)
                else:
                    periode_end = due_date
            else:
                # Invoice berikutnya: mulai dari tanggal 1 bulan due_date
                # Pastikan due_date adalah datetime
                if due_date.day == 1:
                    # Jika jatuh tempo tgl 1, anggap periode layanan adalah bulan SEBELUMNYA full
                    # Contoh: Jatuh tempo 1 Mar, layanan 1-28 Feb
                    periode_end = due_date - timedelta(days=1)
                    periode_start = periode_end.replace(day=1)
                else:
                    # Jika jatuh tempo tgl 28, layanan 1-28 bulan itu
                    periode_start = due_date.replace(day=1)
                    periode_end = due_date

            # Cek apakah ini invoice gabungan
            if float(db_invoice.total_harga or 0) > (harga_normal_full + 1):
                # INI TAGIHAN GABUNGAN
                # Periode pertama (prorate): dari invoice_date sampai akhir bulan itu
                periode_prorate_end = invoice_date.replace(day=1) + relativedelta(months=1, days=-1)
                periode_prorate_str = safe_format_date(periode_prorate_end, "%B %Y")

                # Periode kedua (bulan penuh): bulan berikutnya
                periode_berikutnya_start = periode_prorate_end + relativedelta(days=1)
                periode_berikutnya_end = periode_berikutnya_start + relativedelta(months=1, days=-1)

                deskripsi_xendit = (
                    f"Biaya internet up to {paket.kecepatan} Mbps. "
                    f"Periode {invoice_date.day}-{periode_prorate_end.day} {periode_prorate_str} + "
                    f"Periode {safe_format_date(periode_berikutnya_end, '%B %Y')}"
                )
            else:
                # INI TAGIHAN PRORATE BIASA
                # Generate deskripsi yang cerdas menangani perbedaan bulan/tahun
                if periode_start.month == periode_end.month and periode_start.year == periode_end.year:
                    # Case 1: Dalam satu bulan (e.g. 6-28 Februari 2026)
                    periode_str = safe_format_date(periode_end, "%B %Y")
                    period_desc = f"Periode Tgl {periode_start.day}-{periode_end.day} {periode_str}"
                elif periode_start.year == periode_end.year:
                     # Case 2: Beda bulan, tahun sama (e.g. 6 Februari - 1 Maret 2026)
                     start_month = safe_format_date(periode_start, "%B")
                     end_month_year = safe_format_date(periode_end, "%B %Y")
                     period_desc = f"Periode Tgl {periode_start.day} {start_month} - {periode_end.day} {end_month_year}"
                else:
                     # Case 3: Beda tahun
                     start_full = safe_format_date(periode_start, "%d %B %Y")
                     end_full = safe_format_date(periode_end, "%d %B %Y")
                     period_desc = f"Periode Tgl {start_full} - {end_full}"

                deskripsi_xendit = (
                    f"Biaya berlangganan internet up to {paket.kecepatan} Mbps, "
                    f"{period_desc}"
                )

        else:  # Otomatis
            deskripsi_xendit = (
                f"Biaya berlangganan internet up to {paket.kecepatan} Mbps "
                f"jatuh tempo pembayaran tanggal {jatuh_tempo_str_lengkap}"
            )

        # Format nomor telepon untuk Xendit (tanpa '+')
        no_telp_bersih = ""
        if pelanggan.no_telp:
            # Gunakan helper universal yang handle semua format: 08xx, 62xx, +62xx, 8xx
            no_telp_bersih = normalize_phone_for_xendit(pelanggan.no_telp)

        no_telp_xendit = no_telp_bersih if no_telp_bersih else None

        # Kirim deskripsi yang sudah dinamis ke Xendit dengan rate limiting
        # Determine priority based on customer type
        priority = InvoicePriority.NORMAL
        if hasattr(pelanggan, 'is_vip') and getattr(pelanggan, 'is_vip', False):
            priority = InvoicePriority.HIGH
        elif hasattr(pelanggan, 'tipe') and getattr(pelanggan, 'tipe', '') == 'bulk':
            priority = InvoicePriority.LOW

        xendit_response = await create_invoice_with_rate_limit(
            invoice=db_invoice,
            pelanggan=pelanggan,
            paket=paket,
            deskripsi_xendit=deskripsi_xendit,
            pajak=pajak,
            no_telp_xendit=no_telp_xendit or "",
            priority=priority
        )

        db_invoice.payment_link = xendit_response.get("short_url", xendit_response.get("invoice_url"))
        db_invoice.xendit_id = xendit_response.get("id")
        db_invoice.xendit_external_id = xendit_response.get("external_id")
        await db.commit()
        await db.refresh(db_invoice)
    except Exception as e:
        await db.rollback()
        logger.error(f"Gagal membuat invoice di Xendit: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Gagal membuat invoice di Xendit: {str(e)}")

    # FIX: Eager load pelanggan sebelum return untuk menghindari MissingGreenlet error
    # saat FastAPI serialize response
    stmt_reload = (
        select(InvoiceModel)
        .where(InvoiceModel.id == db_invoice.id)
        .options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.langganan).joinedload(LanggananModel.paket_layanan),
                joinedload(PelangganModel.data_teknis),
            )
        )
    )
    result_reload = await db.execute(stmt_reload)
    db_invoice_loaded = result_reload.unique().scalar_one()

    return db_invoice_loaded


# POST /invoices/{invoice_id}/mark-as-paid - Tandai invoice lunas manual
# Buat nandai invoice udah dibayar secara manual (bukan via Xendit)
# Path parameters:
# - invoice_id: ID invoice yang mau ditandai lunas
# Request body: metode_pembayaran (cash, transfer, dll)
# Response: data invoice yang udah diupdate
# Permission: butuh permission "edit_invoices"
# Fitur:
# - Update status invoice jadi "Lunas"
# - Update tanggal pembayaran
# - Update status langganan jadi "Aktif"
# - Hitung tanggal jatuh tempo berikutnya
# - Trigger re-aktivasi Mikrotik kalo sebelumnya suspended
# - Kirim notifikasi ke Admin/NOC/Finance
# Validation:
# - Cek invoice harus ada
# - Cek invoice belum lunas sebelumnya
# Error handling: 404 kalo invoice nggak ada, 400 kalo udah lunas
@router.post(
    "/{invoice_id}/mark-as-paid",
    response_model=InvoiceSchema,
    dependencies=[Depends(has_permission("edit_invoices"))],
)
async def mark_invoice_as_paid(invoice_id: int, payload: MarkAsPaidRequest, db: AsyncSession = Depends(get_db)):
    """Menandai sebuah invoice sebagai lunas secara manual."""
    # PERBAIKAN: Eager load semua relasi yang dibutuhkan oleh _process_successful_payment
    # untuk menghindari N+1 query dan membuat proses lebih efisien.
    stmt = (
        select(InvoiceModel)
        .where(InvoiceModel.id == invoice_id)
        .options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.langganan).joinedload(LanggananModel.paket_layanan),
                joinedload(PelangganModel.data_teknis),
            )
        )
    )
    invoice = (await db.execute(stmt)).unique().scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice tidak ditemukan")

    if invoice.status_invoice == "Lunas":
        raise HTTPException(status_code=400, detail="Invoice ini sudah lunas.")

    invoice.metode_pembayaran = payload.metode_pembayaran

    # SAFETY NET: Simpan status original untuk potensi Mikrotik rollback
    _original_status = None
    _rollback_langganan = None
    _rollback_data_teknis = None
    if invoice.pelanggan and invoice.pelanggan.langganan:
        _rollback_langganan = invoice.pelanggan.langganan[0]
        _original_status = _rollback_langganan.status
        _rollback_data_teknis = getattr(invoice.pelanggan, 'data_teknis', None)

    await _process_successful_payment(db, invoice)

    try:
        await db.commit()
        await db.refresh(invoice)
    except Exception as commit_error:
        await db.rollback()
        logger.error(f"❌ Commit gagal saat mark-as-paid untuk invoice {invoice_id}: {commit_error}")
        # Rollback Mikrotik jika unsuspend sudah dilakukan
        if _original_status == "Suspended" and _rollback_langganan and _rollback_data_teknis:
            try:
                _rollback_langganan.status = "Suspended"
                await mikrotik_service.trigger_mikrotik_update(
                    db, _rollback_langganan, _rollback_data_teknis, _rollback_data_teknis.id_pelanggan
                )
                logger.info(f"✅ Mikrotik rollback successful for langganan ID {_rollback_langganan.id}")
            except Exception as rollback_error:
                logger.error(f"❌ CRITICAL: Mikrotik rollback FAILED: {rollback_error}")
                _rollback_data_teknis.mikrotik_sync_pending = True
                try:
                    db.add(_rollback_data_teknis)
                    await db.commit()
                except Exception:
                    pass
        raise HTTPException(status_code=500, detail="Gagal menyimpan perubahan pembayaran.")

    # Clear sidebar cache untuk update badge count
    try:
        from .dashboard import clear_sidebar_cache
        clear_sidebar_cache()
    except ImportError:
        pass  # Fallback jika import gagal

    logger.info(f"Invoice {invoice.invoice_number} ditandai lunas secara manual via {payload.metode_pembayaran}")
    
    return invoice


@router.get("/missing-payment-links", response_model=List[InvoiceSchema])
async def get_invoices_missing_payment_links(
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user),
    _: None = Depends(has_permission("view_invoices")),
):
    """
    Get semua invoice yang tidak memiliki payment link dari Xendit.
    Ini berguna untuk finance team agar bisa retry pembuatan payment link.
    """
    stmt = select(InvoiceModel).where(
        InvoiceModel.payment_link.is_(None),
        InvoiceModel.status_invoice == "Belum Dibayar"
    ).order_by(InvoiceModel.tgl_invoice.desc())

    result = await db.execute(stmt)
    invoices = result.scalars().unique().all()

    logger.info(f"Found {len(invoices)} invoices missing payment links")
    return invoices


@router.post("/{invoice_id}/retry-xendit", response_model=dict)
async def retry_invoice_xendit(
    invoice_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user),
    _: None = Depends(has_permission("edit_invoices")),
):
    """
    Manual retry pembuatan payment link Xendit untuk invoice yang gagal.
    Endpoint ini untuk admin agar bisa retry invoice yang gagal secara manual.

    Parameters:
    - invoice_id: ID invoice yang mau di-retry

    Returns:
    - success: status retry
    - message: pesan hasilnya
    - payment_link: link pembayaran jika berhasil

    Process:
    1. Cek invoice harus ada dan belum punya payment link
    2. Reset retry count ke 0
    3. Coba buat payment link lagi ke Xendit
    4. Update invoice jika berhasil
    """
    from math import floor

    # Load invoice dengan semua relasi yang dibutuhkan
    stmt = (
        select(InvoiceModel)
        .where(InvoiceModel.id == invoice_id)
        .options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.langganan).joinedload(LanggananModel.paket_layanan),
                joinedload(PelangganModel.data_teknis),
            )
        )
    )
    invoice = (await db.execute(stmt)).unique().scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice tidak ditemukan")

    if invoice.xendit_id and invoice.payment_link:
        raise HTTPException(status_code=400, detail="Invoice ini sudah punya payment link")

    if invoice.status_invoice == "Lunas":
        raise HTTPException(status_code=400, detail="Invoice sudah lunas")

    try:
        pelanggan = invoice.pelanggan
        if not pelanggan or not pelanggan.langganan:
            raise HTTPException(status_code=400, detail="Data pelanggan/langganan tidak lengkap")

        paket = pelanggan.langganan[0].paket_layanan
        brand = pelanggan.harga_layanan
        data_teknis = pelanggan.data_teknis

        if not all([paket, brand, data_teknis]):
            raise HTTPException(status_code=400, detail="Data paket/brand/data teknis tidak lengkap")

        # Reset retry count untuk manual retry
        invoice.xendit_retry_count = 0
        invoice.xendit_status = "processing"
        invoice.xendit_last_retry = datetime.now()
        await db.flush()

        # Generate deskripsi yang sama seperti generate_single_invoice
        # Convert SQLAlchemy Date ke Python 
        try:
            jatuh_tempo_date = date.fromisoformat(str(invoice.tgl_jatuh_tempo))
            invoice_date = date.fromisoformat(str(invoice.tgl_invoice))
        except (ValueError, TypeError):
            jatuh_tempo_date = date.today()
            invoice_date = date.today()

        jatuh_tempo_str = datetime.combine(jatuh_tempo_date, datetime.min.time()).strftime("%d/%m/%Y")

        if pelanggan.langganan[0].metode_pembayaran == "Prorate":
            # Handle Smart Date Formatting (Consistency with generate_manual_invoice)
            periode_start = invoice_date
            if jatuh_tempo_date.day == 1:
                periode_end = jatuh_tempo_date - timedelta(days=1)
            else:
                periode_end = jatuh_tempo_date

            if periode_start.month == periode_end.month and periode_start.year == periode_end.year:
                period_desc = f"Periode Tgl {periode_start.day}-{periode_end.day} {safe_format_date(periode_end, '%B %Y')}"
            elif periode_start.year == periode_end.year:
                period_desc = f"Periode Tgl {periode_start.day} {safe_format_date(periode_start, '%B')} - {periode_end.day} {safe_format_date(periode_end, '%B %Y')}"
            else:
                period_desc = f"Periode Tgl {safe_format_date(periode_start, '%d %B %Y')} - {safe_format_date(periode_end, '%d %B %Y')}"

            deskripsi_xendit = (
                f"Biaya berlangganan internet up to {paket.kecepatan} Mbps, "
                f"{period_desc}"
            )
        else:  # Otomatis
            deskripsi_xendit = (
                f"Biaya berlangganan internet up to {paket.kecepatan} Mbps "
                f"jatuh tempo pembayaran tanggal {jatuh_tempo_str}"
            )

        # Hitung pajak
        pajak_persen = float(brand.pajak)
        harga_dasar = float(paket.harga)
        pajak = floor(harga_dasar * (pajak_persen / 100) + 0.5)

        no_telp_xendit = normalize_phone_for_xendit(pelanggan.no_telp)

        # Coba buat payment link lagi
        xendit_response = await create_invoice_with_rate_limit(
            invoice=invoice,
            pelanggan=pelanggan,
            paket=paket,
            deskripsi_xendit=deskripsi_xendit,
            pajak=pajak,
            no_telp_xendit=no_telp_xendit,
            priority=InvoicePriority.HIGH  # High priority untuk manual retry
        )

        # Validasi response
        if not xendit_response or not xendit_response.get("id"):
            raise ValueError(f"Invalid Xendit response: {xendit_response}")

        # Update invoice dengan payment link
        invoice.payment_link = xendit_response.get("short_url", xendit_response.get("invoice_url"))
        invoice.xendit_id = xendit_response.get("id")
        invoice.xendit_external_id = xendit_response.get("external_id")
        invoice.xendit_status = "completed"
        invoice.xendit_error_message = None

        await db.commit()

        logger.info(f"✅ Manual retry SUCCESS: Invoice {invoice.invoice_number} - {pelanggan.nama}")
        logger.info(f"📱 Payment link: {invoice.payment_link}")

        return {
            "success": True,
            "message": f"Payment link berhasil dibuat untuk invoice {invoice.invoice_number}",
            "payment_link": invoice.payment_link,
            "xendit_id": invoice.xendit_id,
            "pelanggan": pelanggan.nama
        }

    except Exception as e:
        await db.rollback()

        # Update error tracking
        invoice.xendit_retry_count += 1
        invoice.xendit_status = "failed"
        invoice.xendit_error_message = str(e)
        await db.commit()

        logger.error(f"❌ Manual retry FAILED: Invoice {invoice.invoice_number} - {str(e)}")

        raise HTTPException(
            status_code=500,
            detail=f"Gagal membuat payment link: {str(e)}"
        )


@router.post("/batch-retry-xendit", response_model=dict)
async def batch_retry_failed_invoices(
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user),
    _: None = Depends(has_permission("edit_invoices")),
):
    """
    Batch retry untuk semua invoice yang gagal (belum ada payment link).
    Endpoint ini untuk admin agar bisa retry semua invoice gagal sekaligus.

    Returns:
    - total_processed: total invoice yang diproses
    - success_count: jumlah yang berhasil
    - failed_count: jumlah yang gagal
    - results: detail hasil per invoice
    """
    from math import floor

    # Cari semua invoice yang belum ada payment link-nya
    stmt = (
        select(InvoiceModel)
        .where(
            InvoiceModel.payment_link.is_(None),
            InvoiceModel.status_invoice == "Belum Dibayar"
        )
        .options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.harga_layanan),
                joinedload(PelangganModel.langganan).joinedload(LanggananModel.paket_layanan),
                joinedload(PelangganModel.data_teknis),
            )
        )
        .order_by(InvoiceModel.created_at.desc())
        .limit(50)  # Batasi max 50 invoice per batch
    )

    invoices = (await db.execute(stmt)).unique().scalars().all()

    if not invoices:
        return {
            "success": True,
            "message": "Tidak ada invoice yang perlu di-retry",
            "total_processed": 0,
            "success_count": 0,
            "failed_count": 0,
            "results": []
        }

    total_processed = len(invoices)
    success_count = 0
    failed_count = 0
    results = []

    logger.info(f"🔄 Starting batch retry for {total_processed} invoices by user {current_user.name}")

    for invoice in invoices:
        try:
            pelanggan = invoice.pelanggan
            if not pelanggan or not pelanggan.langganan:
                results.append({
                    "invoice_id": invoice.id,
                    "invoice_number": invoice.invoice_number,
                    "success": False,
                    "message": "Data pelanggan/langganan tidak lengkap"
                })
                failed_count += 1
                continue

            paket = pelanggan.langganan[0].paket_layanan
            brand = pelanggan.harga_layanan

            if not all([paket, brand]):
                results.append({
                    "invoice_id": invoice.id,
                    "invoice_number": invoice.invoice_number,
                    "success": False,
                    "message": "Data paket/brand tidak lengkap"
                })
                failed_count += 1
                continue

            # Reset retry count
            invoice.xendit_retry_count = 0
            invoice.xendit_status = "processing"
            invoice.xendit_last_retry = datetime.now()
            await db.flush()

            # Generate deskripsi
            # Convert SQLAlchemy Date ke Python date dengan aman
            try:
                jatuh_tempo_date = date.fromisoformat(str(invoice.tgl_jatuh_tempo))
                invoice_date = date.fromisoformat(str(invoice.tgl_invoice))
            except (ValueError, TypeError):
                jatuh_tempo_date = date.today()
                invoice_date = date.today()

            jatuh_tempo_str = datetime.combine(jatuh_tempo_date, datetime.min.time()).strftime("%d/%m/%Y")

            if pelanggan.langganan[0].metode_pembayaran == "Prorate":
                # Handle Smart Date Formatting (Consistency with generate_manual_invoice)
                periode_start = invoice_date
                if jatuh_tempo_date.day == 1:
                    periode_end = jatuh_tempo_date - timedelta(days=1)
                else:
                    periode_end = jatuh_tempo_date

                if periode_start.month == periode_end.month and periode_start.year == periode_end.year:
                    period_desc = f"Periode Tgl {periode_start.day}-{periode_end.day} {safe_format_date(periode_end, '%B %Y')}"
                elif periode_start.year == periode_end.year:
                    period_desc = f"Periode Tgl {periode_start.day} {safe_format_date(periode_start, '%B')} - {periode_end.day} {safe_format_date(periode_end, '%B %Y')}"
                else:
                    period_desc = f"Periode Tgl {safe_format_date(periode_start, '%d %B %Y')} - {safe_format_date(periode_end, '%d %B %Y')}"

                deskripsi_xendit = (
                    f"Biaya berlangganan internet up to {paket.kecepatan} Mbps, "
                    f"{period_desc}"
                )
            else:
                deskripsi_xendit = (
                    f"Biaya berlangganan internet up to {paket.kecepatan} Mbps "
                    f"jatuh tempo pembayaran tanggal {jatuh_tempo_str}"
                )

            # Hitung pajak
            pajak_persen = float(brand.pajak)
            harga_dasar = float(paket.harga)
            pajak = floor(harga_dasar * (pajak_persen / 100) + 0.5)

            no_telp_xendit = normalize_phone_for_xendit(pelanggan.no_telp)

            # Coba buat payment link
            xendit_response = await create_invoice_with_rate_limit(
                invoice=invoice,
                pelanggan=pelanggan,
                paket=paket,
                deskripsi_xendit=deskripsi_xendit,
                pajak=pajak,
                no_telp_xendit=no_telp_xendit,
                priority=InvoicePriority.NORMAL
            )

            if xendit_response and xendit_response.get("id"):
                # Update invoice
                invoice.payment_link = xendit_response.get("short_url", xendit_response.get("invoice_url"))
                invoice.xendit_id = xendit_response.get("id")
                invoice.xendit_external_id = xendit_response.get("external_id")
                invoice.xendit_status = "completed"
                invoice.xendit_error_message = None

                results.append({
                    "invoice_id": invoice.id,
                    "invoice_number": invoice.invoice_number,
                    "success": True,
                    "message": "Payment link berhasil dibuat",
                    "payment_link": invoice.payment_link,
                    "pelanggan": pelanggan.nama
                })
                success_count += 1

                logger.info(f"✅ Batch retry SUCCESS: {invoice.invoice_number} - {pelanggan.nama}")
            else:
                raise ValueError("Invalid Xendit response")

        except Exception as e:
            # Update error tracking
            invoice.xendit_retry_count += 1
            invoice.xendit_status = "failed"
            invoice.xendit_error_message = str(e)

            results.append({
                "invoice_id": invoice.id,
                "invoice_number": invoice.invoice_number,
                "success": False,
                "message": str(e)
            })
            failed_count += 1

            logger.error(f"❌ Batch retry FAILED: {invoice.invoice_number} - {str(e)}")

    await db.commit()

    logger.info(f"🏁 Batch retry completed: {success_count} success, {failed_count} failed")

    return {
        "success": True,
        "message": f"Batch retry selesai: {success_count} berhasil, {failed_count} gagal",
        "total_processed": total_processed,
        "success_count": success_count,
        "failed_count": failed_count,
        "results": results
    }


# DELETE /invoices/{invoice_id} - Hapus invoice
# Buat hapus invoice dari sistem
# Path parameters:
# - invoice_id: ID invoice yang mau dihapus
# Response: 204 No Content (sukses tapi nggak ada response body)
# Permission: butuh permission "delete_invoices"
# Warning: HATI-HATI! Ini akan hapus invoice permanen
# Note: Payment link di Xendit mungkin masih aktif
# Error handling: 404 kalau invoice nggak ketemu
@router.delete(
    "/{invoice_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(has_permission("delete_invoices"))],
)
async def delete_invoice(invoice_id: int, db: AsyncSession = Depends(get_db)):
    """Menghapus satu invoice berdasarkan ID-nya."""

    db_invoice = await db.get(InvoiceModel, invoice_id)

    if not db_invoice:
        raise HTTPException(status_code=404, detail="Invoice tidak ditemukan")

    await db.delete(db_invoice)
    await db.commit()

    return None


# Logic atau EndPoint untuk melihat Status Invoice itu masih layar atau sudah kadaluarsa ??


# POST /invoices/internal/update-overdue-status - Update status invoice kadaluarsa (Internal)
# Endpoint internal buat update status invoice yang telat bayar dan suspend layanan
# Endpoint ini HARUS dipanggil oleh scheduler (cron job) setiap hari
# Response: jumlah invoice yang diupdate + jumlah layanan yang disuspend
# Security: include_in_schema=False (nggak muncul di dokumentasi API publik)
# Logic:
# - Cari invoice "Belum Dibayar" yang udah lewat 5 hari dari jatuh tempo
# - Update status jadi "Kadaluarsa"
# - Update status langganan jadi "Suspended"
# - Trigger suspend Mikrotik
# Aturan: Kadaluarsa kalo hari ini adalah hari ke-6 setelah jatuh tempo (lewat 5 hari)
# Use case: automated system untuk penagihan dan penonaktifan layanan
# Error handling: graceful error handling, log semua error
@router.post(
    "/internal/update-overdue-status",
    status_code=status.HTTP_200_OK,
    include_in_schema=False,
)
async def update_overdue_invoices(db: AsyncSession = Depends(get_db)):
    """
    Endpoint internal untuk memperbarui status invoice menjadi 'Kadaluarsa' dan men-suspend layanan.
    Endpoint ini HARUS dipanggil oleh scheduler (cron job) setiap hari dan tidak boleh terekspos ke publik.
    `include_in_schema=False` digunakan agar tidak muncul di dokumentasi API publik.
    """
    today = date.today()
    # Aturan: Kadaluarsa jika hari ini adalah hari ke-6 setelah jatuh tempo (lewat 5 hari).
    overdue_threshold_date = today - timedelta(days=5)

    logger = logging.getLogger("app.routers.invoice")
    logger.info("Scheduler-triggered job: Mencari invoice kadaluarsa...")

    # 1. Cari semua invoice yang 'Belum Dibayar' dan sudah melewati masa tenggang
    stmt = (
        select(InvoiceModel)
        # PERBAIKAN: Eager load data_teknis untuk mencegah N+1 query di dalam loop.
        .options(
            joinedload(InvoiceModel.pelanggan).options(
                joinedload(PelangganModel.langganan),
                joinedload(PelangganModel.data_teknis),
            )
        ).where(
            and_(
                InvoiceModel.status_invoice == "Belum Dibayar",
                InvoiceModel.tgl_jatuh_tempo < overdue_threshold_date,
            )
        )
    )

    # FIX: Tambahkan .unique() untuk collection eager loading
    overdue_invoices = (await db.execute(stmt)).scalars().unique().all()

    if not overdue_invoices:
        logger.info("Tidak ada invoice kadaluarsa yang perlu diperbarui.")
        return {"message": "Tidak ada invoice kadaluarsa yang perlu diperbarui."}

    updated_count = 0
    suspended_count = 0
    # 2. Update status dan suspend layanan untuk setiap invoice yang ditemukan
    for invoice in overdue_invoices:
        invoice.status_invoice = "Kadaluarsa"
        db.add(invoice)
        updated_count += 1

        # Logika untuk men-suspend layanan pelanggan
        try:
            pelanggan = invoice.pelanggan
            if pelanggan and pelanggan.langganan:
                langganan_pelanggan = pelanggan.langganan[0]
                if langganan_pelanggan.status != "Suspended":
                    langganan_pelanggan.status = "Suspended"
                    db.add(langganan_pelanggan)

                    # PERBAIKAN: Panggil trigger_mikrotik_update dengan argumen yang benar.
                    data_teknis = pelanggan.data_teknis
                    if data_teknis:
                        await mikrotik_service.trigger_mikrotik_update(
                            db,
                            langganan_pelanggan,
                            data_teknis,
                            data_teknis.id_pelanggan,
                        )
                        suspended_count += 1
                        logger.info(f"Layanan untuk {pelanggan.nama} (Invoice: {invoice.invoice_number}) telah di-suspend.")
                    else:
                        logger.error(
                            f"Data Teknis tidak ditemukan untuk pelanggan {pelanggan.nama}, suspend di Mikrotik dilewati."
                        )
        except Exception as e:
            logger.error(f"Gagal men-suspend layanan untuk invoice {invoice.invoice_number}: {e}")

    await db.commit()

    message = f"Proses selesai. {updated_count} invoice diperbarui menjadi 'Kadaluarsa'. {suspended_count} layanan di-suspend."
    logger.info(message)
    return {"message": message}


# GET /invoices/export-payment-links-excel - Export payment links ke Excel
# Buat export semua payment link invoice ke file Excel
# Query parameters:
# - search: filter pencarian (sama seperti di list)
# - status_invoice: filter berdasarkan status
# - start_date: filter tanggal mulai
# - end_date: filter tanggal akhir
# Response: file Excel (.xlsx) dengan kolom:
#   ID Invoice, Nomor Invoice, Nama Pelanggan, ID Pelanggan, Alamat,
#   Total Harga, Status Invoice, Tanggal Invoice, Tanggal Jatuh Tempo,
#   Payment Link, Email, No. Telepon, Brand
# Use case: buat share payment links ke tim collection atau customer
# Performance: query dengan eager loading biar efficient
# Format: Excel dengan header bold dan auto-size columns
@router.get("/export-payment-links-excel")
async def export_payment_links_excel(
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user),  # PROTECTED
    search: Optional[str] = None,
    status_invoice: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
):
    """Export payment links dari invoice ke file Excel."""
    from datetime import datetime

    query = (
        select(InvoiceModel)
        .join(InvoiceModel.pelanggan)
        .options(
            joinedload(InvoiceModel.pelanggan)
            .joinedload(PelangganModel.harga_layanan)
        )
        # Eager load langganan untuk metode_pembayaran
        .options(
            joinedload(InvoiceModel.pelanggan)
            .joinedload(PelangganModel.langganan)
        )
    )

    if search:
        search_term = f"%{search}%"
        # Cek apakah search term adalah angka (untuk pelanggan_id)
        is_numeric_search = search.strip().isdigit()

        query = query.where(
            or_(
                InvoiceModel.invoice_number.ilike(search_term),
                PelangganModel.nama.ilike(search_term),
                InvoiceModel.id_pelanggan.ilike(search_term),
                # Tambahkan filter langsung untuk pelanggan_id (BigInteger foreign key)
                # Ini memperbaiki bug di mana riwayat pembayaran di view langganan tidak muncul
                InvoiceModel.pelanggan_id == int(search.strip()) if is_numeric_search else False,
            )
        )

    if status_invoice:
        query = query.where(InvoiceModel.status_invoice == status_invoice)

    if start_date:
        query = query.where(InvoiceModel.tgl_jatuh_tempo >= start_date)
    if end_date:
        query = query.where(InvoiceModel.tgl_jatuh_tempo <= end_date)

    # Hanya ambil invoice dengan payment_link
    query = query.where(InvoiceModel.payment_link.isnot(None))

    result = await db.execute(query)
    # FIX: Tambahkan .unique() untuk collection eager loading
    invoices = result.scalars().unique().all()

    # Query untuk mendapatkan invoice pertama (ID terkecil) setiap pelanggan
    # Ini untuk menandai "NEW USER"
    from sqlalchemy import func as sql_func
    pelanggan_ids = list(set([inv.pelanggan_id for inv in invoices]))
    first_invoice_map = {}

    if pelanggan_ids:
        first_invoice_query = (
            select(
                InvoiceModel.pelanggan_id,
                sql_func.min(InvoiceModel.id).label('first_invoice_id')
            )
            .where(InvoiceModel.pelanggan_id.in_(pelanggan_ids))
            .group_by(InvoiceModel.pelanggan_id)
        )
        first_invoice_result = await db.execute(first_invoice_query)
        for row in first_invoice_result:
            first_invoice_map[row.pelanggan_id] = row.first_invoice_id

    # Buat workbook dan worksheet pertama untuk Payment Links
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Payment Links Invoice"

    # Definisikan header
    headers = [
        "ID Invoice",
        "Nomor Invoice",
        "Nama Pelanggan",
        "ID Pelanggan",
        "Alamat Pelanggan",
        "Total Harga",
        "Status Invoice",
        "Tipe Invoice",  # Kolom baru untuk jenis invoice
        "Tanggal Invoice",
        "Tanggal Jatuh Tempo",
        "TANGGAL BAYAR",  # Kolom baru untuk tanggal pembayaran
        "Payment Link",
        "Email",
        "No. Telepon",
        "Brand",
        "NEW USER",  # Kolom baru untuk menandai user baru
    ]

    # Menambahkan header ke worksheet (dengan null check)
    if ws is not None:
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            # Gunakan import langsung untuk styles
            from openpyxl.styles import Font, PatternFill, Alignment

            if cell is not None:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Isi data (dengan null check)
    if ws is not None:
        for row_num, invoice in enumerate(invoices, 2):
            ws.cell(row=row_num, column=1, value=invoice.id)
            ws.cell(row=row_num, column=2, value=invoice.invoice_number)
            ws.cell(row=row_num, column=3, value=invoice.pelanggan.nama if invoice.pelanggan else "")
            ws.cell(row=row_num, column=4, value=invoice.id_pelanggan)
            ws.cell(row=row_num, column=5, value=invoice.pelanggan.alamat if invoice.pelanggan else "")
            ws.cell(row=row_num, column=6, value=float(invoice.total_harga) if invoice.total_harga else 0)
            ws.cell(row=row_num, column=7, value=invoice.status_invoice)

            # Tipe invoice dengan mapping yang user-friendly
            # Format: [Tipe] - [Metode] - [Diskon jika ada]
            # Contoh: "Otomatis - Diskon", "Prorate - Reinvoice", "Manual"

            # Tentukan metode pembayaran (ambil dari langganan pelanggan)
            metode_pembayaran = "Otomatis"  # Default
            if invoice.pelanggan and hasattr(invoice.pelanggan, 'langganan'):
                # Cari langganan yang aktif
                langganan_aktif = None
                for langganan in invoice.pelanggan.langganan:
                    if langganan.status == "Aktif":
                        langganan_aktif = langganan
                        break
                if langganan_aktif:
                    metode_pembayaran = langganan_aktif.metode_pembayaran or "Otomatis"

            # Bangun tipe invoice string
            type_parts = []

            # Tambah tipe invoice (Otomatis/Manual)
            if getattr(invoice, 'is_reinvoice', False):
                type_parts.append("Reinvoice")
                # Untuk reinvoice, gunakan metode pembayaran dari langganan
                type_parts.append(metode_pembayaran)
            else:
                invoice_type = invoice.invoice_type or "manual"
                if invoice_type == "automatic":
                    type_parts.append("Otomatis")
                elif invoice_type == "manual":
                    type_parts.append("Manual")
                else:
                    type_parts.append(invoice_type.capitalize())

                # Tambah metode pembayaran jika berbeda
                if metode_pembayaran == "Prorate":
                    type_parts.append("Prorate")

            # Tambah flag diskon jika ada
            if getattr(invoice, 'diskon_id', None) and invoice.diskon_id:
                type_parts.append("Diskon")

            # Gabungkan semua bagian dengan " - "
            invoice_type_display = " - ".join(type_parts)
            ws.cell(row=row_num, column=8, value=invoice_type_display)

            # Handle SQLAlchemy Date untuk Excel export
            invoice_date = invoice.tgl_invoice
            due_date = invoice.tgl_jatuh_tempo

            ws.cell(row=row_num, column=9, value=safe_format_date(invoice_date, "%Y-%m-%d"))
            ws.cell(row=row_num, column=10, value=safe_format_date(due_date, "%Y-%m-%d"))

            # TANGGAL BAYAR - Tampilkan paid_at jika status Lunas
            if invoice.status_invoice == "Lunas" and invoice.paid_at:
                # Format paid_at ke datetime WIB untuk display
                paid_at_wib = invoice.paid_at
                if paid_at_wib:
                    ws.cell(row=row_num, column=11, value=paid_at_wib.strftime("%Y-%m-%d %H:%M"))
                else:
                    ws.cell(row=row_num, column=11, value="-")
            else:
                ws.cell(row=row_num, column=11, value="-")

            ws.cell(row=row_num, column=12, value=invoice.payment_link)
            ws.cell(row=row_num, column=13, value=invoice.email or "")
            ws.cell(row=row_num, column=14, value=invoice.no_telp or "")
            ws.cell(row=row_num, column=15, value=invoice.brand or "")

            # NEW USER - Tandai jika ini adalah invoice pertama untuk pelanggan
            is_new_user = first_invoice_map.get(invoice.pelanggan_id) == invoice.id
            ws.cell(row=row_num, column=16, value="NEW USER" if is_new_user else "")

        # Auto-adjust column width (dengan null check)
        from openpyxl.utils import get_column_letter

        if ws is not None:
            for column in ws.columns:
                max_length = 0
                # Handle column[0].column yang mungkin None
                first_cell = column[0] if column else None
                if first_cell and hasattr(first_cell, "column") and first_cell.column is not None:
                    column_letter = get_column_letter(first_cell.column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    if ws is not None:
                        ws.column_dimensions[column_letter].width = adjusted_width

    # Buat sheet kedua untuk Matrix Persentase Invoice
    ws_matrix = wb.create_sheet("Matrix Persentase Invoice")
    if ws_matrix is not None:
        # Hitung statistik dari invoices yang sudah difilter
        total_invoices = len(invoices)
        if total_invoices > 0:
            # Hitung berdasarkan status invoice
            lunas_count = len([inv for inv in invoices if inv.status_invoice == 'Lunas'])
            belum_dibayar_count = len([inv for inv in invoices if inv.status_invoice == 'Belum Dibayar'])
            kadaluarsa_count = len([inv for inv in invoices if inv.status_invoice == 'Kadaluarsa'])

            # Hitung berdasarkan tipe invoice
            otomatis_count = len([inv for inv in invoices if inv.invoice_type == 'automatic'])
            manual_count = len([inv for inv in invoices if inv.invoice_type == 'manual'])
            reinvoice_count = len([inv for inv in invoices if getattr(inv, 'is_reinvoice', False)])

            # Hitung persentase status
            lunas_percent = (lunas_count / total_invoices) * 100
            belum_dibayar_percent = (belum_dibayar_count / total_invoices) * 100
            kadaluarsa_percent = (kadaluarsa_count / total_invoices) * 100

            # Hitung persentase tipe
            otomatis_percent = (otomatis_count / total_invoices) * 100
            manual_percent = (manual_count / total_invoices) * 100
            reinvoice_percent = (reinvoice_count / total_invoices) * 100

            # Format tanggal untuk header
            start_date_str = start_date.strftime('%d/%m/%Y') if start_date else 'Awal'
            end_date_str = end_date.strftime('%d/%m/%Y') if end_date else 'Sekarang'

            # Styling untuk header matrix
            header_font = Font(bold=True, size=14)
            title_font = Font(bold=True, size=12)
            data_font = Font(bold=False, size=11)
            blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            green_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
            red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

            # Header Matrix
        ws_matrix.merge_cells('A1:D1')
        ws_matrix.cell(row=1, column=1, value="MATRIX LAPORAN PERSENTASE INVOICE")
        ws_matrix.cell(row=1, column=1).font = header_font
        ws_matrix.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        # Informasi Periode dan Total
        ws_matrix.cell(row=2, column=1, value=f"Periode: {start_date_str} - {end_date_str}")
        ws_matrix.cell(row=2, column=1).font = title_font
        ws_matrix.merge_cells('A2:D2')

        ws_matrix.cell(row=3, column=1, value=f"Total Invoice (Filter Aktif): {total_invoices}")
        ws_matrix.cell(row=3, column=1).font = title_font
        ws_matrix.merge_cells('A3:D3')

        # Spasi
        ws_matrix.cell(row=4, column=1, value="")

        # Header Tabel
        ws_matrix.cell(row=5, column=1, value="Status Invoice")
        ws_matrix.cell(row=5, column=2, value="Jumlah")
        ws_matrix.cell(row=5, column=3, value="Persentase")
        ws_matrix.cell(row=5, column=4, value="Visual")

        # Styling header tabel
        for col in range(1, 5):
            cell = ws_matrix.cell(row=5, column=col)
            cell.font = title_font
            cell.fill = gray_fill
            cell.alignment = Alignment(horizontal='center')

        # Data Total Invoice
        ws_matrix.cell(row=6, column=1, value="Total Invoice")
        ws_matrix.cell(row=6, column=2, value=total_invoices)
        ws_matrix.cell(row=6, column=3, value="100.0%")
        ws_matrix.cell(row=6, column=4, value="████████████████████")
        for col in range(1, 5):
            cell = ws_matrix.cell(row=6, column=col)
            cell.font = data_font
            cell.fill = blue_fill

        # Data Invoice Lunas
        ws_matrix.cell(row=7, column=1, value="Invoice Lunas")
        ws_matrix.cell(row=7, column=2, value=lunas_count)
        ws_matrix.cell(row=7, column=3, value=f"{lunas_percent:.1f}%")
        # Visual bar chart dengan text blocks
        bar_length = int(lunas_percent / 5)  # 1 block = 5%
        ws_matrix.cell(row=7, column=4, value="█" * bar_length)
        for col in range(1, 4):
            cell = ws_matrix.cell(row=7, column=col)
            cell.font = data_font
            cell.fill = green_fill

        # Data Invoice Belum Dibayar
        ws_matrix.cell(row=8, column=1, value="Invoice Belum Dibayar")
        ws_matrix.cell(row=8, column=2, value=belum_dibayar_count)
        ws_matrix.cell(row=8, column=3, value=f"{belum_dibayar_percent:.1f}%")
        bar_length = int(belum_dibayar_percent / 5)
        ws_matrix.cell(row=8, column=4, value="█" * bar_length)
        for col in range(1, 4):
            cell = ws_matrix.cell(row=8, column=col)
            cell.font = data_font
            cell.fill = yellow_fill

        # Data Invoice Kadaluarsa
        ws_matrix.cell(row=9, column=1, value="Invoice Kadaluarsa")
        ws_matrix.cell(row=9, column=2, value=kadaluarsa_count)
        ws_matrix.cell(row=9, column=3, value=f"{kadaluarsa_percent:.1f}%")
        bar_length = int(kadaluarsa_percent / 5)
        ws_matrix.cell(row=9, column=4, value="█" * bar_length)
        for col in range(1, 4):
            cell = ws_matrix.cell(row=9, column=col)
            cell.font = data_font
            cell.fill = red_fill

        # Spasi
        ws_matrix.cell(row=10, column=1, value="")

        # Header Tipe Invoice
        ws_matrix.cell(row=11, column=1, value="TIPE INVOICE")
        ws_matrix.cell(row=11, column=1).font = title_font
        ws_matrix.merge_cells('A11:D11')

        # Data Invoice Otomatis
        ws_matrix.cell(row=12, column=1, value="Invoice Otomatis")
        ws_matrix.cell(row=12, column=2, value=otomatis_count)
        ws_matrix.cell(row=12, column=3, value=f"{otomatis_percent:.1f}%")
        bar_length = int(otomatis_percent / 5)
        ws_matrix.cell(row=12, column=4, value="█" * bar_length)
        for col in range(1, 4):
            cell = ws_matrix.cell(row=12, column=col)
            cell.font = data_font
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Blue

        # Data Invoice Manual
        ws_matrix.cell(row=13, column=1, value="Invoice Manual")
        ws_matrix.cell(row=13, column=2, value=manual_count)
        ws_matrix.cell(row=13, column=3, value=f"{manual_percent:.1f}%")
        bar_length = int(manual_percent / 5)
        ws_matrix.cell(row=13, column=4, value="█" * bar_length)
        for col in range(1, 4):
            cell = ws_matrix.cell(row=13, column=col)
            cell.font = data_font
            cell.fill = PatternFill(start_color="F3E6FF", end_color="F3E6FF", fill_type="solid")  # Purple

        # Data Reinvoice
        ws_matrix.cell(row=14, column=1, value="Reinvoice")
        ws_matrix.cell(row=14, column=2, value=reinvoice_count)
        ws_matrix.cell(row=14, column=3, value=f"{reinvoice_percent:.1f}%")
        bar_length = int(reinvoice_percent / 5)
        ws_matrix.cell(row=14, column=4, value="█" * bar_length)
        for col in range(1, 4):
            cell = ws_matrix.cell(row=14, column=col)
            cell.font = data_font
            cell.fill = PatternFill(start_color="FFE6F3", end_color="FFE6F3", fill_type="solid")  # Pink

        # Spasi
        ws_matrix.cell(row=15, column=1, value="")

        # Insight/Kesimpulan
        ws_matrix.cell(row=16, column=1, value="INSIGHT & KESIMPULAN")
        ws_matrix.cell(row=16, column=1).font = title_font
        ws_matrix.merge_cells('A16:D16')

        # Analisis pembayaran
        if lunas_percent >= 70:
            payment_insight = f"Tingkat pembayaran sangat baik ({lunas_percent:.1f}% lunas)"
        elif lunas_percent >= 50:
            payment_insight = f"Tingkat pembayaran cukup baik ({lunas_percent:.1f}% lunas)"
        else:
            payment_insight = f"Tingkat pembayaran perlu ditingkatkan ({lunas_percent:.1f}% lunas)"

        ws_matrix.cell(row=17, column=1, value=payment_insight)
        ws_matrix.merge_cells('A17:D17')

        # Analisis tipe invoice
        if otomatis_percent > 50:
            type_insight = f"Sistem invoice otomatis efektif ({otomatis_percent:.1f}%)"
        elif manual_percent > 50:
            type_insight = f"Masih dominan invoice manual ({manual_percent:.1f}%)"
        elif reinvoice_percent > 20:
            type_insight = f"Tingkat reinvoice tinggi ({reinvoice_percent:.1f}%) - perlu evaluasi"
        else:
            type_insight = f"Distribusi tipe invoice normal"

        ws_matrix.cell(row=18, column=1, value=type_insight)
        ws_matrix.merge_cells('A18:D18')

        # Rekomendasi tindakan
        if kadaluarsa_percent > 20:
            rekomendasi = "Perlu follow-up intensif untuk invoice kadaluarsa"
        elif belum_dibayar_percent > 50:
            rekomendasi = "Perlu reminder rutin untuk pembayaran"
        elif reinvoice_percent > 20:
            rekomendasi = "Evaluasi penyebab tingginya reinvoice dan optimalkan reminder"
        else:
            rekomendasi = "Status pembayaran dalam kondisi normal"

        ws_matrix.cell(row=19, column=1, value=rekomendasi)
        ws_matrix.merge_cells('A19:D19')

        # Auto-adjust column width untuk sheet matrix
        for column in ws_matrix.columns:
            max_length = 0
            first_cell = column[0] if column else None
            if first_cell and hasattr(first_cell, "column") and first_cell.column is not None:
                column_letter = get_column_letter(first_cell.column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_matrix.column_dimensions[column_letter].width = adjusted_width

    else:
        # Jika tidak ada data, gunakan styling dasar
        header_font = Font(bold=True, size=14)
        ws_matrix.cell(row=1, column=1, value="MATRIX LAPORAN PERSENTASE INVOICE")
        ws_matrix.cell(row=1, column=1).font = header_font
        ws_matrix.merge_cells('A1:D1')

        ws_matrix.cell(row=2, column=1, value="Tidak ada data invoice yang memenuhi filter")
        ws_matrix.merge_cells('A2:D2')

    # Buat sheet ketiga untuk Suspended User Matrix
    try:
        # Get suspended matrix data
        matrix_data = await get_suspended_invoice_matrix(db, None, None)

        ws_suspended = wb.create_sheet("Suspended User Matrix")

        # Styling untuk suspended matrix
        blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        green_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
        red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        # Header
        ws_suspended.cell(row=1, column=1, value="MATRIX SUSPENDED USERS VS INVOICE STATUS")
        ws_suspended.cell(row=1, column=1).font = header_font
        ws_suspended.merge_cells('A1:E1')

        ws_suspended.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        ws_suspended.merge_cells('A2:E2')

        # Section 1: Users Summary
        ws_suspended.cell(row=4, column=1, value="USERS SUMMARY")
        ws_suspended.cell(row=4, column=1).font = title_font
        ws_suspended.merge_cells('A4:E4')

        users_headers = ["Metric", "Count", "Percentage"]
        for col, header in enumerate(users_headers, 1):
            ws_suspended.cell(row=5, column=col, value=header)
            ws_suspended.cell(row=5, column=col).fill = gray_fill

        # Users Data
        users_summary = matrix_data['users_summary']
        ws_suspended.cell(row=6, column=1, value="Total All Users")
        ws_suspended.cell(row=6, column=2, value=users_summary['total_all_users'])
        ws_suspended.cell(row=6, column=3, value="100.0%")
        for col in range(1, 4):
            ws_suspended.cell(row=6, column=col).fill = blue_fill

        ws_suspended.cell(row=7, column=1, value="Total Aktif Users")
        ws_suspended.cell(row=7, column=2, value=users_summary['total_aktif_users'])
        ws_suspended.cell(row=7, column=3, value=f"{(users_summary['total_aktif_users']/users_summary['total_all_users']*100):.1f}%" if users_summary['total_all_users'] > 0 else "0%")
        for col in range(1, 4):
            ws_suspended.cell(row=7, column=col).fill = green_fill

        ws_suspended.cell(row=8, column=1, value="Total Suspended Users")
        ws_suspended.cell(row=8, column=2, value=users_summary['total_suspended_users'])
        ws_suspended.cell(row=8, column=3, value=f"{users_summary['suspended_percentage']}%")
        for col in range(1, 4):
            ws_suspended.cell(row=8, column=col).fill = red_fill

        # Section 2: Invoice Status Matrix
        ws_suspended.cell(row=10, column=1, value="INVOICE STATUS MATRIX")
        ws_suspended.cell(row=10, column=1).font = title_font
        ws_suspended.merge_cells('A10:E10')

        matrix_headers = ["User Status", "Total Invoices", "Lunas", "Belum Dibayar", "Bermasalah (Exp/Kad)"]
        for col, header in enumerate(matrix_headers, 1):
            ws_suspended.cell(row=11, column=col, value=header)
            ws_suspended.cell(row=11, column=col).fill = gray_fill

        # Aktif Users Invoice Data
        aktif_inv = matrix_data['aktif_users_invoices']
        ws_suspended.cell(row=12, column=1, value="Aktif Users")
        ws_suspended.cell(row=12, column=2, value=aktif_inv['total'])
        ws_suspended.cell(row=12, column=3, value=aktif_inv['lunas'])
        ws_suspended.cell(row=12, column=4, value=aktif_inv['belum_dibayar'])
        ws_suspended.cell(row=12, column=5, value=aktif_inv['expired'] + aktif_inv['kadaluarsa'])
        for col in range(1, 6):
            ws_suspended.cell(row=12, column=col).fill = green_fill

        # Suspended Users Invoice Data
        susp_inv = matrix_data['suspended_users_invoices']
        ws_suspended.cell(row=13, column=1, value="Suspended Users")
        ws_suspended.cell(row=13, column=2, value=susp_inv['total'])
        ws_suspended.cell(row=13, column=3, value=susp_inv['lunas'])
        ws_suspended.cell(row=13, column=4, value=susp_inv['belum_dibayar'])
        ws_suspended.cell(row=13, column=5, value=susp_inv['expired'] + susp_inv['kadaluarsa'])
        for col in range(1, 6):
            ws_suspended.cell(row=13, column=col).fill = red_fill

        # Section 3: Suspended Analysis
        ws_suspended.cell(row=15, column=1, value="SUSPENDED USERS ANALYSIS")
        ws_suspended.cell(row=15, column=1).font = title_font
        ws_suspended.merge_cells('A15:E15')

        analysis_headers = ["Analysis Type", "Count", "Rate", "Status"]
        for col, header in enumerate(analysis_headers, 1):
            ws_suspended.cell(row=16, column=col, value=header)
            ws_suspended.cell(row=16, column=col).fill = gray_fill

        susp_analysis = matrix_data['suspended_analysis']

        ws_suspended.cell(row=17, column=1, value="Suspended with Unpaid")
        ws_suspended.cell(row=17, column=2, value=susp_analysis['suspended_with_unpaid_invoices'])
        ws_suspended.cell(row=17, column=3, value=f"{susp_analysis['unpaid_collection_rate']}%")
        ws_suspended.cell(row=17, column=4, value="Critical" if susp_analysis['unpaid_collection_rate'] > 50 else "Monitor")
        for col in range(1, 5):
            ws_suspended.cell(row=17, column=col).fill = red_fill if susp_analysis['unpaid_collection_rate'] > 50 else yellow_fill

        ws_suspended.cell(row=18, column=1, value="Suspended with Paid")
        ws_suspended.cell(row=18, column=2, value=susp_analysis['suspended_with_paid_invoices'])
        ws_suspended.cell(row=18, column=3, value=f"{susp_analysis['critical_recovery_rate']}%")
        ws_suspended.cell(row=18, column=4, value="Good")
        for col in range(1, 5):
            ws_suspended.cell(row=18, column=col).fill = green_fill

        ws_suspended.cell(row=19, column=1, value="Suspended No Invoice")
        ws_suspended.cell(row=19, column=2, value=susp_analysis['suspended_without_invoices'])
        ws_suspended.cell(row=19, column=3, value="N/A")
        ws_suspended.cell(row=19, column=4, value="New/Closed")
        for col in range(1, 5):
            ws_suspended.cell(row=19, column=col).fill = blue_fill

        # Section 4: Insights
        ws_suspended.cell(row=21, column=1, value="INSIGHTS & REKOMENDASI")
        ws_suspended.cell(row=21, column=1).font = title_font
        ws_suspended.merge_cells('A21:E21')

        insights = matrix_data['insights']
        ws_suspended.cell(row=22, column=1, value=f"Suspension Rate: {insights['suspension_rate']}")
        ws_suspended.merge_cells('A22:E22')

        ws_suspended.cell(row=23, column=1, value=f"Collection Challenge: {insights['collection_challenge']}")
        ws_suspended.merge_cells('A23:E23')

        ws_suspended.cell(row=24, column=1, value=f"Recovery Potential: {insights['recovery_potential']}")
        ws_suspended.merge_cells('A24:E24')

        # Auto-adjust columns untuk sheet suspended
        for column in ws_suspended.columns:
            max_length = 0
            first_cell = column[0] if column else None
            if first_cell and hasattr(first_cell, 'column') and first_cell.column is not None:
                column_letter = get_column_letter(first_cell.column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_suspended.column_dimensions[column_letter].width = adjusted_width

    except Exception as e:
        # If error getting suspended matrix, create empty sheet
        ws_suspended = wb.create_sheet("Suspended User Matrix")
        ws_suspended.cell(row=1, column=1, value="SUSPENDED USER MATRIX")
        ws_suspended.cell(row=1, column=1).font = header_font
        ws_suspended.cell(row=2, column=1, value="Error generating matrix data")
        ws_suspended.merge_cells('A2:E2')

    # Simpan workbook ke BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Kembalikan file sebagai response
    return Response(
        buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=invoice_report_comprehensive.xlsx"},
    )


# GET /invoices/count - Hitung total jumlah invoice
# Buat ngambil total jumlah invoice di database dengan optional filter
# Query parameters:
# - status_invoice: filter berdasarkan status (Belum Dibayar, Lunas, Kadaluarsa)
# - start_date: filter tanggal jatuh tempo mulai dari
# - end_date: filter tanggal jatuh tempo sampai dengan
# - search: cari berdasarkan nomor invoice, nama pelanggan, atau ID pelanggan
# Response: integer (total count)
# Use case: buat dashboard atau statistik
# Performance: simple count query dengan filter yang sama seperti list endpoint
@router.get("/count", response_model=int)
async def get_invoice_count(
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user),  # PROTECTED
    search: Optional[str] = None,
    status_invoice: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    exclude_expired: Optional[bool] = False,
):
    """
    Menghitung total jumlah invoice dengan filter opsional (termasuk ARSIP).
    """
    # 1. Hitung dari tabel invoice utama
    count_query = select(func.count(InvoiceModel.id))
    count_query = count_query.join(InvoiceModel.pelanggan)

    if search:
        search_term = f"%{search}%"
        # Cek apakah search term adalah angka (untuk pelanggan_id)
        is_numeric_search = search.strip().isdigit()

        count_query = count_query.where(
            or_(
                InvoiceModel.invoice_number.ilike(search_term),
                PelangganModel.nama.ilike(search_term),
                InvoiceModel.id_pelanggan.ilike(search_term),
                # Tambahkan filter langsung untuk pelanggan_id (BigInteger foreign key)
                # Ini memperbaiki bug di mana riwayat pembayaran di view langganan tidak muncul
                InvoiceModel.pelanggan_id == int(search.strip()) if is_numeric_search else False,
            )
        )

    if status_invoice:
        count_query = count_query.where(InvoiceModel.status_invoice == status_invoice)

    if start_date:
        count_query = count_query.where(InvoiceModel.tgl_jatuh_tempo >= start_date)
    if end_date:
        count_query = count_query.where(InvoiceModel.tgl_jatuh_tempo <= end_date)

    if exclude_expired:
        from datetime import date
        overdue_threshold_date = date.today() - timedelta(days=5)
        count_query = count_query.where(
            or_(
                InvoiceModel.status_invoice == 'Lunas',
                and_(
                    InvoiceModel.status_invoice == 'Belum Dibayar',
                    InvoiceModel.tgl_jatuh_tempo >= overdue_threshold_date
                )
            )
        )

    result = await db.execute(count_query)
    main_count = result.scalar_one()

    # 2. Hitung dari tabel invoice archive
    count_query_archive = select(func.count(InvoiceArchiveModel.id))
    count_query_archive = count_query_archive.join(InvoiceArchiveModel.pelanggan)

    if search:
        search_term = f"%{search}%"
        count_query_archive = count_query_archive.where(
            or_(
                InvoiceArchiveModel.invoice_number.ilike(search_term),
                PelangganModel.nama.ilike(search_term),
                InvoiceArchiveModel.id_pelanggan.ilike(search_term),
            )
        )

    if status_invoice:
        count_query_archive = count_query_archive.where(InvoiceArchiveModel.status_invoice == status_invoice)

    if start_date:
        count_query_archive = count_query_archive.where(InvoiceArchiveModel.tgl_jatuh_tempo >= start_date)
    if end_date:
        count_query_archive = count_query_archive.where(InvoiceArchiveModel.tgl_jatuh_tempo <= end_date)

    if exclude_expired:
        from datetime import date
        overdue_threshold_date = date.today() - timedelta(days=5)
        count_query_archive = count_query_archive.where(
            or_(
                InvoiceArchiveModel.status_invoice == 'Lunas',
                and_(
                    InvoiceArchiveModel.status_invoice == 'Belum Dibayar',
                    InvoiceArchiveModel.tgl_jatuh_tempo >= overdue_threshold_date
                )
            )
        )

    result_archive = await db.execute(count_query_archive)
    archive_count = result_archive.scalar_one()

    return main_count + archive_count


# ====================================================================
# MONITORING & FIX ENDPOINTS - PRODUCTION RELIABILITY
# ====================================================================

@router.get("/generation-status")
async def get_invoice_generation_status(
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user)
):
    """Simple dashboard untuk monitor invoice generation"""

    today = date.today()

    # Total invoice hari ini
    total_today = await db.execute(
        select(func.count(InvoiceModel.id))
        .where(InvoiceModel.tgl_invoice == today)
    )

    # Invoice dengan payment link
    with_payment_link = await db.execute(
        select(func.count(InvoiceModel.id))
        .where(
            InvoiceModel.tgl_invoice == today,
            InvoiceModel.payment_link.isnot(None)
        )
    )

    total = total_today.scalar() or 0
    success = with_payment_link.scalar() or 0
    failed = total - success

    return {
        "date": today.isoformat(),
        "total_invoices": total,
        "successful_payments": success,
        "failed_payments": failed,
        "success_rate": round((success / total * 100) if total > 0 else 0, 1),
        "status": "HEALTHY" if failed == 0 else "NEEDS_ATTENTION" if failed <= 3 else "CRITICAL"
    }


@router.post("/fix-missing-payment-links")
async def fix_missing_payment_links(
    limit: int = 20,
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user)
):
    """Fix invoices yang tidak punya payment link"""

    from ..services.rate_limiter import create_invoice_with_rate_limit

    # Cari invoice tanpa payment link
    invoices_to_fix = await db.execute(
        select(InvoiceModel)
        .where(InvoiceModel.payment_link.is_(None))
        .order_by(InvoiceModel.created_at.desc())
        .limit(limit)
        .options(
            selectinload(InvoiceModel.pelanggan).selectinload(PelangganModel.harga_layanan),
            selectinload(InvoiceModel.pelanggan).selectinload(PelangganModel.langganan).selectinload(LanggananModel.paket_layanan),
        )
    )

    fixed_count = 0
    failed_count = 0

    for invoice in invoices_to_fix.scalars().all():
        try:
            pelanggan = invoice.pelanggan
            if not pelanggan or not pelanggan.langganan:
                continue

            paket = pelanggan.langganan[0].paket_layanan

            # Generate payment link
            xendit_response = await create_invoice_with_rate_limit(
                invoice=invoice,
                pelanggan=pelanggan,
                paket=paket,
                deskripsi_xendit=f"Invoice Payment - {invoice.invoice_number}",
                pajak=float(invoice.total_harga) - float(paket.harga) if paket else 0,
                no_telp_xendit=normalize_phone_for_xendit(pelanggan.no_telp)
            )

            # Update invoice
            invoice.payment_link = xendit_response.get("short_url", xendit_response.get("invoice_url"))
            invoice.xendit_id = xendit_response.get("id")
            invoice.xendit_external_id = xendit_response.get("external_id")

            fixed_count += 1

        except Exception as e:
            failed_count += 1
            logger.error(f"Failed to fix invoice {invoice.invoice_number}: {e}")

    await db.commit()

    return {
        "message": f"Fixed {fixed_count} invoices, {failed_count} failed",
        "fixed_count": fixed_count,
        "failed_count": failed_count
    }


@router.get("/affected-phone-format")
async def get_affected_phone_format_invoices(
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user),
    _: None = Depends(has_permission("view_invoices")),
):
    """
    Endpoint untuk mengidentifikasi invoice yang terdampak format nomor telepon salah.
    
    Masalah: Nomor yang disimpan di DB sebagai 62xxx (misalnya 6281283725103)
    sebelumnya diformat menjadi +6262xxx (dobel prefix) saat dikirim ke Xendit,
    sehingga pelanggan tersebut TIDAK menerima WhatsApp notification dari Xendit.
    
    Endpoint ini mengembalikan:
    - Daftar invoice "Belum Dibayar" yang nomornya berawalan 62 di database
    - Payment link yang bisa dishare manual via WhatsApp oleh finance
    - Nomor WhatsApp yang benar (sudah dinormalisasi)
    
    USAGE: GET /invoices/affected-phone-format
    Finance kemudian bisa copy payment link dan kirim manual ke pelanggan.
    """
    from ..utils.phone_utils import normalize_phone_for_xendit

    try:
        # Cari invoice yang masih "Belum Dibayar" dan punya payment link
        stmt = (
            select(InvoiceModel)
            .where(
                InvoiceModel.status_invoice == "Belum Dibayar",
                InvoiceModel.payment_link.isnot(None),
            )
            .options(
                joinedload(InvoiceModel.pelanggan)
            )
            .order_by(InvoiceModel.created_at.desc())
        )

        result = await db.execute(stmt)
        invoices = result.unique().scalars().all()

        affected_invoices = []
        for inv in invoices:
            if not inv.pelanggan or not inv.pelanggan.no_telp:
                continue

            raw_phone = inv.pelanggan.no_telp.strip()
            
            # Identifikasi nomor yang PASTI terdampak:
            # 1. Nomor yang sudah dalam format 62xxx (tanpa 0) di database
            #    → sebelumnya jadi +6262xxx (SALAH)
            # 2. Nomor yang dalam format +62xxx di database  
            #    → sebelumnya jadi +6262xxx (SALAH)
            digits_only = ''.join(c for c in raw_phone if c.isdigit())
            
            is_affected = False
            reason = ""
            
            if digits_only.startswith('62') and not raw_phone.startswith('0'):
                is_affected = True
                reason = f"Nomor di DB: '{raw_phone}' → dikirim ke Xendit sebagai '+62{digits_only}' (dobel 62)"
            
            if is_affected:
                correct_phone = normalize_phone_for_xendit(raw_phone)
                affected_invoices.append({
                    "invoice_id": inv.id,
                    "invoice_number": inv.invoice_number,
                    "pelanggan_nama": inv.pelanggan.nama,
                    "no_telp_di_db": raw_phone,
                    "no_telp_salah_lama": f"+62{digits_only}",
                    "no_telp_benar": correct_phone,
                    "no_wa_share": correct_phone.replace("+", ""),
                    "payment_link": inv.payment_link,
                    "total_harga": float(inv.total_harga) if inv.total_harga else 0,
                    "tgl_jatuh_tempo": str(inv.tgl_jatuh_tempo),
                    "reason": reason,
                })

        return {
            "success": True,
            "total_affected": len(affected_invoices),
            "message": (
                f"Ditemukan {len(affected_invoices)} invoice yang terdampak format nomor telepon salah. "
                f"Finance bisa share payment link ke nomor WhatsApp yang benar secara manual."
                if affected_invoices
                else "Tidak ada invoice yang terdampak. Semua notifikasi WhatsApp sudah terkirim dengan benar."
            ),
            "affected_invoices": affected_invoices,
        }

    except Exception as e:
        logger.error(f"Error checking affected phone format: {e}")
        raise HTTPException(status_code=500, detail=f"Gagal mengecek invoice terdampak: {str(e)}")

@router.get("/summary")
async def get_invoice_summary(
    current_user: UserModel = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """
    Endpoint untuk mendapatkan summary dashboard invoice:
    - Total User (jumlah pelanggan aktif)
    - Total Invoice
    - Jumlah PAID
    - Jumlah Expired
    - Jumlah Reinvoice (invoice yang tergenerate otomatis & manual)
    """
    try:
        # Query Total User (pelanggan aktif)
        # Pelanggan model tidak punya deleted_at, jadi query semua
        total_users_query = select(func.count(PelangganModel.id))
        total_users_result = await db.execute(total_users_query)
        total_users = total_users_result.scalar()

        # Sederhanakan query - hanya dari tabel invoices utama dulu
        # Query Total Invoice
        total_invoice_query = select(func.count()).where(InvoiceModel.deleted_at.is_(None))
        total_invoice_result = await db.execute(total_invoice_query)
        total_invoices = total_invoice_result.scalar() or 0

        # Query Jumlah PAID invoice
        paid_query = select(func.count()).where(
            and_(
                InvoiceModel.status_invoice == "Lunas",
                InvoiceModel.deleted_at.is_(None)
            )
        )
        paid_result = await db.execute(paid_query)
        total_paid = paid_result.scalar() or 0

        # Query Jumlah Expired invoice (sederhana)
        expired_query = select(func.count()).where(
            and_(
                InvoiceModel.status_invoice == "Expired",
                InvoiceModel.deleted_at.is_(None)
            )
        )
        expired_result = await db.execute(expired_query)
        total_expired = expired_result.scalar() or 0

        # Query Jumlah Reinvoice
        reinvoice_query = select(func.count()).where(
            and_(
                InvoiceModel.is_reinvoice == True,
                InvoiceModel.deleted_at.is_(None)
            )
        )
        reinvoice_result = await db.execute(reinvoice_query)
        total_reinvoice = reinvoice_result.scalar() or 0

        # Query Invoice Otomatis
        automatic_query = select(func.count()).where(
            and_(
                InvoiceModel.invoice_type == "automatic",
                InvoiceModel.deleted_at.is_(None)
            )
        )
        automatic_result = await db.execute(automatic_query)
        total_automatic = automatic_result.scalar() or 0

        # Query Invoice Manual
        manual_query = select(func.count()).where(
            and_(
                InvoiceModel.invoice_type == "manual",
                InvoiceModel.deleted_at.is_(None)
            )
        )
        manual_result = await db.execute(manual_query)
        total_manual = manual_result.scalar() or 0

        # Summary data
        summary_data = {
            "total_users": total_users or 0,
            "total_invoices": total_invoices,
            "total_paid": total_paid,
            "total_expired": total_expired,
            "total_reinvoice": total_reinvoice,
            "invoice_types": {
                "automatic": total_automatic,
                "manual": total_manual,
                "reinvoice": total_reinvoice
            },
            "last_updated": datetime.now().isoformat()
        }

        return summary_data

    except Exception as e:
        logger.error(f"Error getting invoice summary: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to get invoice summary"
        )


@router.get("/suspended-invoice-matrix")
async def get_suspended_invoice_matrix(
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user),
    _: None = Depends(has_permission("view_invoices")),
):
    """
    Summary pencocokan user yang suspended dengan status invoice:
    - Total Users (aktif, suspended, total)
    - Status Invoice breakdown per user status
    - Matrix analysis untuk identifikasi pattern
    """
    try:
        # Query total users by status
        total_users_query = select(func.count(PelangganModel.id))
        total_users_result = await db.execute(total_users_query)
        total_all_users = total_users_result.scalar() or 0

        # Query aktif users
        aktif_users_query = select(func.count(LanggananModel.id)).where(LanggananModel.status == "Aktif")
        aktif_users_result = await db.execute(aktif_users_query)
        total_aktif_users = aktif_users_result.scalar() or 0

        # Query suspended users
        suspended_users_query = select(func.count(LanggananModel.id)).where(LanggananModel.status == "Suspended")
        suspended_users_result = await db.execute(suspended_users_query)
        total_suspended_users = suspended_users_result.scalar() or 0

        # Query invoice status for aktif users
        aktif_invoices_query = select(
            InvoiceModel.status_invoice,
            func.count(InvoiceModel.id).label('jumlah')
        ).join(
            InvoiceModel.pelanggan
        ).join(
            PelangganModel.langganan
        ).where(
            and_(
                LanggananModel.status == "Aktif",
                InvoiceModel.deleted_at.is_(None)
            )
        ).group_by(InvoiceModel.status_invoice)

        aktif_invoices_result = await db.execute(aktif_invoices_query)
        aktif_invoice_stats = {
            row.status_invoice: row.jumlah
            for row in aktif_invoices_result
        }

        # Query invoice status for suspended users
        suspended_invoices_query = select(
            InvoiceModel.status_invoice,
            func.count(InvoiceModel.id).label('jumlah')
        ).join(
            InvoiceModel.pelanggan
        ).join(
            PelangganModel.langganan
        ).where(
            and_(
                LanggananModel.status == "Suspended",
                InvoiceModel.deleted_at.is_(None)
            )
        ).group_by(InvoiceModel.status_invoice)

        suspended_invoices_result = await db.execute(suspended_invoices_query)
        suspended_invoice_stats = {
            row.status_invoice: row.jumlah
            for row in suspended_invoices_result
        }

        # Query suspended users dengan unpaid invoices (critical cases)
        suspended_unpaid_query = select(
            func.count(func.distinct(PelangganModel.id)).label('count')
        ).join(
            PelangganModel.langganan
        ).join(
            PelangganModel.invoices
        ).where(
            and_(
                LanggananModel.status == "Suspended",
                InvoiceModel.status_invoice.in_(['Belum Dibayar', 'Expired', 'Kadaluarsa']),
                InvoiceModel.deleted_at.is_(None)
            )
        )

        suspended_unpaid_result = await db.execute(suspended_unpaid_query)
        suspended_with_unpaid = suspended_unpaid_result.scalar() or 0

        # Query suspended users dengan paid invoices (good cases)
        suspended_paid_query = select(
            func.count(func.distinct(PelangganModel.id)).label('count')
        ).join(
            PelangganModel.langganan
        ).join(
            PelangganModel.invoices
        ).where(
            and_(
                LanggananModel.status == "Suspended",
                InvoiceModel.status_invoice == "Lunas",
                InvoiceModel.deleted_at.is_(None)
            )
        )

        suspended_paid_result = await db.execute(suspended_paid_query)
        suspended_with_paid = suspended_paid_result.scalar() or 0

        # Matrix data
        matrix_data = {
            "users_summary": {
                "total_all_users": total_all_users,
                "total_aktif_users": total_aktif_users,
                "total_suspended_users": total_suspended_users,
                "suspended_percentage": round((total_suspended_users / total_all_users * 100), 2) if total_all_users > 0 else 0
            },
            "aktif_users_invoices": {
                "total": sum(aktif_invoice_stats.values()),
                "lunas": aktif_invoice_stats.get('Lunas', 0),
                "belum_dibayar": aktif_invoice_stats.get('Belum Dibayar', 0),
                "expired": aktif_invoice_stats.get('Expired', 0),
                "kadaluarsa": aktif_invoice_stats.get('Kadaluarsa', 0)
            },
            "suspended_users_invoices": {
                "total": sum(suspended_invoice_stats.values()),
                "lunas": suspended_invoice_stats.get('Lunas', 0),
                "belum_dibayar": suspended_invoice_stats.get('Belum Dibayar', 0),
                "expired": suspended_invoice_stats.get('Expired', 0),
                "kadaluarsa": suspended_invoice_stats.get('Kadaluarsa', 0)
            },
            "suspended_analysis": {
                "suspended_with_unpaid_invoices": suspended_with_unpaid,
                "suspended_with_paid_invoices": suspended_with_paid,
                "suspended_without_invoices": total_suspended_users - (suspended_with_unpaid + suspended_with_paid),
                "critical_recovery_rate": round((suspended_with_paid / total_suspended_users * 100), 2) if total_suspended_users > 0 else 0,
                "unpaid_collection_rate": round((suspended_with_unpaid / total_suspended_users * 100), 2) if total_suspended_users > 0 else 0
            },
            "insights": {
                "suspension_rate": "Tinggi" if (total_suspended_users / total_all_users * 100) > 30 else "Normal",
                "collection_challenge": "Tinggi" if suspended_with_unpaid > suspended_with_paid else "Rendah",
                "recovery_potential": f"{suspended_with_unpaid} users perlu follow-up pembayaran"
            },
            "last_updated": datetime.now().isoformat()
        }

        return matrix_data

    except Exception as e:
        logger.error(f"Error getting suspended invoice matrix: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to get suspended invoice matrix"
        )


# ====================================================================
# INVOICE GENERATION MONITORING
# ====================================================================

@router.get("/skipped-invoice-generation")
async def get_skipped_invoice_generation(
    target_date: Optional[date] = None,
    db: AsyncSession = Depends(get_db),
    current_user: UserModel = Depends(get_current_active_user)
):
    """
    Monitoring untuk mendeteksi pelanggan yang seharusnya dapat invoice otomatis
    tapi terlewat (skipped/missed) saat job generate invoice berjalan.
    
    Logic:
    1. Cari semua langganan aktif dengan tgl_jatuh_tempo = target_date
    2. Cek mana yang sudah punya invoice untuk periode tersebut
    3. Return yang belum punya invoice (terlewat/skipped)
    
    Args:
        target_date: Tanggal jatuh tempo yang mau dicek (default: tanggal 1 bulan depan)
    
    Returns:
        - total_should_have_invoice: Jumlah pelanggan yang seharusnya dapat invoice
        - total_generated: Jumlah invoice yang berhasil di-generate
        - total_skipped: Jumlah pelanggan yang terlewat
        - skipped_customers: Detail pelanggan yang terlewat
    """
    
    # Default target_date = tanggal 1 bulan depan (sesuai logika H-5)
    if not target_date:
        today = date.today()
        # Hitung tanggal 1 bulan depan
        if today.month == 12:
            target_date = date(today.year + 1, 1, 1)
        else:
            target_date = date(today.year, today.month + 1, 1)
    
    # 1. Cari semua langganan aktif dengan tgl_jatuh_tempo = target_date
    should_have_invoice_stmt = (
        select(LanggananModel)
        .options(
            selectinload(LanggananModel.pelanggan).selectinload(PelangganModel.data_teknis),
            selectinload(LanggananModel.pelanggan).selectinload(PelangganModel.harga_layanan),
            selectinload(LanggananModel.paket_layanan)
        )
        .where(
            LanggananModel.tgl_jatuh_tempo == target_date,
            LanggananModel.status == "Aktif"
        )
    )
    
    should_have_invoice_result = await db.execute(should_have_invoice_stmt)
    all_langganan = should_have_invoice_result.scalars().unique().all()
    
    total_should_have = len(all_langganan)
    
    # 2. Cari invoice yang sudah di-generate untuk periode ini
    # Calculate month period for target_date
    target_year = target_date.year
    target_month = target_date.month
    start_of_month = date(target_year, target_month, 1)
    if target_month == 12:
        end_of_month = date(target_year + 1, 1, 1) - timedelta(days=1)
    else:
        end_of_month = date(target_year, target_month + 1, 1) - timedelta(days=1)
    
    # Get all pelanggan_id yang sudah punya invoice di periode ini
    existing_invoices_stmt = (
        select(InvoiceModel.pelanggan_id)
        .where(
            InvoiceModel.tgl_jatuh_tempo.between(start_of_month, end_of_month)
        )
        .distinct()
    )
    
    existing_invoices_result = await db.execute(existing_invoices_stmt)
    pelanggan_with_invoice = set(existing_invoices_result.scalars().all())
    
    # 3. Filter langganan yang belum punya invoice (skipped)
    skipped_customers = []
    for langganan in all_langganan:
        if langganan.pelanggan_id not in pelanggan_with_invoice:
            pelanggan = langganan.pelanggan
            data_teknis = pelanggan.data_teknis if pelanggan else None
            paket = langganan.paket_layanan
            brand = pelanggan.harga_layanan if pelanggan else None
            
            # Detect skip reason
            reasons = []
            if not pelanggan:
                reasons.append("Data pelanggan tidak ditemukan")
            if not paket:
                reasons.append("Data paket layanan tidak ditemukan")
            if not brand:
                reasons.append("Data brand/harga layanan tidak ditemukan")
            if not data_teknis:
                reasons.append("Data teknis tidak ditemukan")
            if pelanggan and not pelanggan.email:
                reasons.append("Email pelanggan kosong")
            if pelanggan and not pelanggan.no_telp:
                reasons.append("Nomor telepon pelanggan kosong")
            if langganan.status != "Aktif":
                reasons.append(f"Status langganan: {langganan.status}")
            if not reasons:
                reasons.append("Kemungkinan error saat generate invoice atau Xendit API gagal")
            
            skipped_customers.append({
                "langganan_id": langganan.id,
                "pelanggan_id": langganan.pelanggan_id,
                "id_pelanggan": data_teknis.id_pelanggan if data_teknis else f"PLG-{langganan.pelanggan_id}",
                "nama": pelanggan.nama if pelanggan else "N/A",
                "alamat": pelanggan.alamat if pelanggan else "N/A",
                "no_telp": pelanggan.no_telp if pelanggan else "N/A",
                "email": pelanggan.email if pelanggan else "N/A",
                "paket": f"{paket.kecepatan} Mbps" if paket else "N/A",
                "brand": brand.brand if brand else "N/A",
                "tgl_jatuh_tempo": target_date.isoformat(),
                "status_langganan": langganan.status,
                "metode_pembayaran": langganan.metode_pembayaran,
                "reason": " | ".join(reasons)
            })
    
    total_generated = total_should_have - len(skipped_customers)
    
    return {
        "target_date": target_date.isoformat(),
        "summary": {
            "total_should_have_invoice": total_should_have,
            "total_generated": total_generated,
            "total_skipped": len(skipped_customers),
            "success_rate": round((total_generated / total_should_have * 100) if total_should_have > 0 else 100, 1),
            "status": "HEALTHY" if len(skipped_customers) == 0 else "NEEDS_ATTENTION" if len(skipped_customers) <= 5 else "CRITICAL"
        },
        "skipped_customers": skipped_customers,
        "generated_at": datetime.now().isoformat()
    }


# GET /invoices/income-stats - Estimasi Income untuk Finance Team
# Endpoint untuk menampilkan statistik pendapatan dari invoice dan reinvoice
# Query parameters:
# - period: filter periode (this_month, last_month, this_year, custom)
# - start_date: tanggal mulai untuk custom period
# - end_date: tanggal akhir untuk custom period
# Response: total invoice, total reinvoice, count, dan persentase
@router.get("/income-stats")
async def get_income_stats(
    db: AsyncSession = Depends(get_db),
    period: str = "this_month",  # this_month, last_month, this_year, custom
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
):
    """
    Mendapatkan statistik pendapatan dari invoice dan reinvoice.
    Berguna untuk finance team melihat estimasi income.
    """
    from datetime import date
    from dateutil.relativedelta import relativedelta
    
    # Tentukan range tanggal berdasarkan period
    today = date.today()
    
    if period == "this_month":
        start = date(today.year, today.month, 1)
        # Last day of current month
        if today.month == 12:
            end = date(today.year, 12, 31)
        else:
            end = date(today.year, today.month + 1, 1) - timedelta(days=1)
    
    elif period == "last_month":
        # First day of last month
        first_of_this_month = date(today.year, today.month, 1)
        start = first_of_this_month - relativedelta(months=1)
        # Last day of last month
        end = first_of_this_month - timedelta(days=1)
    
    elif period == "this_year":
        start = date(today.year, 1, 1)
        end = date(today.year, 12, 31)
    
    elif period == "custom":
        if not start_date or not end_date:
            raise HTTPException(
                status_code=400,
                detail="start_date dan end_date harus diisi untuk period 'custom'"
            )
        start = start_date
        end = end_date
    
    else:
        raise HTTPException(
            status_code=400,
            detail="Period tidak valid. Pilihan: this_month, last_month, this_year, custom"
        )
    
    # Query untuk regular invoice (is_reinvoice = False)
    regular_invoice_query = select(
        func.sum(InvoiceModel.total_harga).label("total"),
        func.count(InvoiceModel.id).label("count")
    ).where(
        and_(
            InvoiceModel.tgl_invoice >= start,
            InvoiceModel.tgl_invoice <= end,
            InvoiceModel.is_reinvoice == False,
            InvoiceModel.status_invoice != "Dibatalkan"  # Exclude cancelled invoices
        )
    )
    
    result_regular = await db.execute(regular_invoice_query)
    regular_data = result_regular.one()
    
    # Query untuk reinvoice (is_reinvoice = True)
    reinvoice_query = select(
        func.sum(InvoiceModel.total_harga).label("total"),
        func.count(InvoiceModel.id).label("count")
    ).where(
        and_(
            InvoiceModel.tgl_invoice >= start,
            InvoiceModel.tgl_invoice <= end,
            InvoiceModel.is_reinvoice == True,
            InvoiceModel.status_invoice != "Dibatalkan"  # Exclude cancelled invoices
        )
    )
    
    result_reinvoice = await db.execute(reinvoice_query)
    reinvoice_data = result_reinvoice.one()
    
    # Calculate totals
    total_invoice = float(regular_data.total or 0)
    count_invoice = int(regular_data.count or 0)
    
    total_reinvoice = float(reinvoice_data.total or 0)
    count_reinvoice = int(reinvoice_data.count or 0)
    
    grand_total = total_invoice + total_reinvoice
    total_count = count_invoice + count_reinvoice
    
    # Calculate percentages
    invoice_percentage = round((total_invoice / grand_total * 100) if grand_total > 0 else 0, 1)
    reinvoice_percentage = round((total_reinvoice / grand_total * 100) if grand_total > 0 else 0, 1)
    
    return {
        "period": period,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "invoice": {
            "total": total_invoice,
            "count": count_invoice,
            "percentage": invoice_percentage
        },
        "reinvoice": {
            "total": total_reinvoice,
            "count": count_reinvoice,
            "percentage": reinvoice_percentage
        },
        "summary": {
            "grand_total": grand_total,
            "total_count": total_count
        },
        "generated_at": datetime.now().isoformat()
    }
