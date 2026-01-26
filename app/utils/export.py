"""
Centralized export utilities supporting CSV and Excel formats
"""

import io
import csv
from typing import List, Dict, Any, Optional, Union, Callable
from datetime import datetime
from fastapi.responses import StreamingResponse
import logging

# For Excel support
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class BaseExporter:
    """Base class for export functionality"""

    @staticmethod
    def prepare_export_data(
        raw_data: List[Any],
        field_mapping: Optional[Dict[str, str]] = None,
        exclude_fields: Optional[List[str]] = None,
        transform_functions: Optional[Dict[str, Callable]] = None,
    ) -> List[Dict[str, Any]]:
        """
        Prepare data untuk export dengan field mapping dan transformations
        Menghilangkan duplikasi data preparation logic
        """
        processed_data = []

        for item in raw_data:
            # Convert object to dict
            if hasattr(item, "__dict__"):
                item_dict = item.__dict__.copy()
                # Remove SQLAlchemy internal attributes
                item_dict = {k: v for k, v in item_dict.items() if not k.startswith("_")}
            elif isinstance(item, dict):
                item_dict = item.copy()
            else:
                # For simple types or other objects
                item_dict = {"value": item}

            # Apply field mapping
            if field_mapping:
                mapped_dict = {}
                for export_field, source_field in field_mapping.items():
                    mapped_dict[export_field] = item_dict.get(source_field, "")
                item_dict = mapped_dict

            # Exclude fields
            if exclude_fields:
                item_dict = {k: v for k, v in item_dict.items() if k not in exclude_fields}

            # Apply transformation functions
            if transform_functions:
                for field, transform_func in transform_functions.items():
                    if field in item_dict:
                        try:
                            item_dict[field] = transform_func(item_dict[field])
                        except Exception as e:
                            logger.warning(f"Failed to transform field {field}: {e}")
                            item_dict[field] = str(item_dict[field]) if item_dict[field] else ""

            # Handle None values
            for key, value in item_dict.items():
                if value is None:
                    item_dict[key] = ""

            processed_data.append(item_dict)

        return processed_data


class CSVExporter(BaseExporter):
    """
    CSV export utility
    """

    @staticmethod
    def create_csv_response(
        data: List[Dict[str, Any]], filename_prefix: str, headers: Optional[List[str]] = None, include_bom: bool = True
    ) -> StreamingResponse:
        """
        Create CSV response dengan format yang konsisten
        """
        try:
            # Create StringIO untuk menampung CSV data
            output = io.StringIO()

            # Add BOM untuk Excel compatibility jika required
            if include_bom:
                output.write("\ufeff")

            # Determine headers
            if headers is None and data:
                headers = list(data[0].keys())
            elif headers is None:
                headers = []

            # Create CSV writer dengan semicolon (;) untuk compatibilitas Excel di Indonesia
            writer = csv.DictWriter(output, fieldnames=headers, delimiter=";")
            writer.writeheader()

            # Write data rows
            if data:
                writer.writerows(data)

            # Reset string position
            output.seek(0)

            # Generate filename dengan timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{filename_prefix}_{timestamp}.csv"

            # Create response headers
            response_headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

            # Return StreamingResponse
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode("utf-8")),
                headers=response_headers,
                media_type="text/csv; charset=utf-8",
            )

        except Exception as e:
            logger.error(f"Failed to create CSV response for {filename_prefix}: {e}")
            raise

    @staticmethod
    def create_csv_template(headers: List[str], sample_data: List[Dict[str, Any]], filename_prefix: str) -> StreamingResponse:
        """
        Create CSV template untuk import dengan sample data
        """
        try:
            output = io.StringIO()
            output.write("\ufeff")  # BOM untuk Excel

            # Menggunakan semicolon agar langsung rapi di Excel Indonesia
            writer = csv.DictWriter(output, fieldnames=headers, delimiter=";")
            writer.writeheader()

            # Write sample data jika ada
            if sample_data:
                writer.writerows(sample_data)

            output.seek(0)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"template_{filename_prefix}_{timestamp}.csv"

            response_headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

            return StreamingResponse(
                io.BytesIO(output.getvalue().encode("utf-8")),
                headers=response_headers,
                media_type="text/csv; charset=utf-8",
            )

        except Exception as e:
            logger.error(f"Failed to create CSV template for {filename_prefix}: {e}")
            raise


class ExcelExporter(BaseExporter):
    """
    Excel export utility with enhanced formatting
    """

    @staticmethod
    def create_excel_response(
        data: List[Dict[str, Any]],
        filename_prefix: str,
        headers: Optional[List[str]] = None,
        sheet_name: str = "Data",
        auto_column_width: bool = True,
        style_headers: bool = True
    ) -> StreamingResponse:
        """
        Create Excel response dengan formatting yang menarik
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Determine headers
            if headers is None and data:
                headers = list(data[0].keys())
            elif headers is None:
                headers = []

            # Write headers with styling
            if headers:
                header_row = 1
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=header_row, column=col_num, value=header)

                    if style_headers:
                        # Style headers: bold, background color, borders
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin")
                        )

            # Write data rows
            if data:
                start_row = 2  # Start after header
                for row_num, row_data in enumerate(data, start_row):
                    for col_num, header in enumerate(headers, 1):
                        value = row_data.get(header, "")
                        cell = ws.cell(row=row_num, column=col_num, value=value)

                        # Add borders to data cells
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin")
                        )

                        # Auto-adjust row height for better readability
                        ws.row_dimensions[row_num].height = 20

            # Auto-adjust column widths if requested
            if auto_column_width:
                for col_num in range(1, len(headers) + 1):
                    column_letter = get_column_letter(col_num)

                    # Find the maximum length in the column
                    max_length = 0
                    if headers:
                        header_length = len(str(headers[col_num - 1]))
                        max_length = header_length

                    for row in ws.iter_rows(min_row=2, max_col=col_num, max_row=ws.max_row):
                        cell_value = row[col_num - 1].value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))

                    # Set column width with some padding
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 to prevent too wide columns
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze header row
            ws.freeze_panes = "A2"

            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Generate filename dengan timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{filename_prefix}_{timestamp}.xlsx"

            # Create response headers
            response_headers = {
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }

            return StreamingResponse(
                io.BytesIO(output.getvalue()),
                headers=response_headers,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            logger.error(f"Failed to create Excel response for {filename_prefix}: {e}")
            raise

    @staticmethod
    def create_excel_template(headers: List[str], sample_data: List[Dict[str, Any]], filename_prefix: str, sheet_name: str = "Template") -> StreamingResponse:
        """
        Create Excel template untuk import dengan sample data
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Write headers with styling
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )

            # Write sample data if provided
            if sample_data:
                for row_num, row_data in enumerate(sample_data, 2):
                    for col_num, header in enumerate(headers, 1):
                        value = row_data.get(header, "")
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin")
                        )

            # Auto-adjust column widths
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                max_length = len(str(headers[col_num - 1]))

                for row in ws.iter_rows(min_row=2, max_col=col_num, max_row=len(sample_data) + 1):
                    cell_value = row[col_num - 1].value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze header row
            ws.freeze_panes = "A2"

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"template_{filename_prefix}_{timestamp}.xlsx"

            response_headers = {
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }

            return StreamingResponse(
                io.BytesIO(output.getvalue()),
                headers=response_headers,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            logger.error(f"Failed to create Excel template for {filename_prefix}: {e}")
            raise


class ExportManager:
    """
    Unified export manager that handles both CSV and Excel exports
    """

    @staticmethod
    def create_export_response(
        data: List[Dict[str, Any]],
        filename_prefix: str,
        export_format: str = "csv",
        headers: Optional[List[str]] = None,
        **kwargs
    ) -> StreamingResponse:
        """
        Create export response in specified format (csv or excel)
        """
        export_format = export_format.lower()

        if export_format == "csv":
            return CSVExporter.create_csv_response(data, filename_prefix, headers, **kwargs)
        elif export_format in ["excel", "xlsx"]:
            return ExcelExporter.create_excel_response(data, filename_prefix, headers, **kwargs)
        else:
            raise ValueError(f"Unsupported export format: {export_format}. Supported formats: csv, excel, xlsx")

    @staticmethod
    def create_template_response(
        headers: List[str],
        sample_data: List[Dict[str, Any]],
        filename_prefix: str,
        export_format: str = "csv",
        **kwargs
    ) -> StreamingResponse:
        """
        Create template response in specified format (csv or excel)
        """
        export_format = export_format.lower()

        if export_format == "csv":
            return CSVExporter.create_csv_template(headers, sample_data, filename_prefix)
        elif export_format in ["excel", "xlsx"]:
            return ExcelExporter.create_excel_template(headers, sample_data, filename_prefix, **kwargs)
        else:
            raise ValueError(f"Unsupported export format: {export_format}. Supported formats: csv, excel, xlsx")


class CSVImportHelper:
    """
    Helper utilities untuk CSV import operations
    """

    @staticmethod
    def validate_csv_headers(
        expected_headers: List[str], actual_headers: List[str], case_sensitive: bool = False
    ) -> tuple[bool, List[str]]:
        """
        Validate CSV headers terhadap expected headers
        Returns: (is_valid, missing_headers)
        """
        if not case_sensitive:
            expected_headers = [h.lower() for h in expected_headers]
            actual_headers = [h.lower() for h in actual_headers]

        missing_headers = [h for h in expected_headers if h not in actual_headers]
        is_valid = len(missing_headers) == 0

        return is_valid, missing_headers

    @staticmethod
    def normalize_field_names(data_dict: Dict[str, Any], field_mapping: Dict[str, str]) -> Dict[str, Any]:
        """
        Normalize field names berdasarkan mapping
        """
        normalized = {}
        for csv_field, model_field in field_mapping.items():
            if csv_field in data_dict:
                normalized[model_field] = data_dict[csv_field]

        return normalized

    @staticmethod
    def clean_csv_value(value: str) -> str:
        """
        Clean CSV value dari extra whitespace dan unwanted characters
        """
        if value is None:
            return ""

        # Convert to string and strip whitespace
        cleaned = str(value).strip()

        # Remove BOM and other invisible characters
        cleaned = cleaned.replace("\ufeff", "")
        cleaned = cleaned.replace("\u200b", "")

        return cleaned


# Predefined export configurations untuk common use cases
class ExportConfigurations:
    """
    Predefined configurations untuk export scenarios yang umum
    """

    PELANGGAN_EXPORT = {
        "headers": ["ID", "Nama", "Email", "No Telepon", "Alamat", "No KTP", "Tanggal Instalasi", "Brand"],
        "field_mapping": {
            "ID": "id",
            "Nama": "nama",
            "Email": "email",
            "No Telepon": "no_telp",
            "Alamat": "alamat",
            "No KTP": "no_ktp",
            "Tanggal Instalasi": "tanggal_instalasi",
            "Brand": "id_brand",
            # "Status": "status",
        },
        "exclude_fields": ["password", "internal_notes"],
        "transform_functions": {"Tanggal Instalasi": lambda x: str(x) if x else "", "id": lambda x: str(x) if x else ""},
    }

    DATA_TEKNIS_EXPORT = {
        "headers": [
            "ID Pelanggan", "Nama Pelanggan", "Email Pelanggan", "Alamat", "Alamat Lengkap",
            "Nomor Telepon", "IP Pelanggan", "Profile PPPoE", "VLAN", "SN ONT",
            "Nama Mikrotik Server", "Kode ODP", "Port ODP", "OLT Custom",
            "PON", "OTB", "ODC", "ONU Power (dBm)", "Status"
        ],
        "field_mapping": {
            "ID Pelanggan": "id_pelanggan",
            "Nama Pelanggan": "pelanggan_nama",
            "Email Pelanggan": "email_pelanggan",
            "Alamat": "alamat",
            "Alamat Lengkap": "alamat_2",
            "Nomor Telepon": "no_telp",
            "IP Pelanggan": "ip_pelanggan",
            "Profile PPPoE": "profile_pppoe",
            "VLAN": "vlan",
            "SN ONT": "sn",
            "Nama Mikrotik Server": "nama_mikrotik_server",
            "Kode ODP": "kode_odp",
            "Port ODP": "port_odp",
            "OLT Custom": "olt_custom",
            "PON": "pon",
            "OTB": "otb",
            "ODC": "odc",
            "ONU Power (dBm)": "onu_power",
            "Status": "status",
        },
        "exclude_fields": ["internal_config", "secrets"],
    }

    INVOICE_EXPORT = {
        "headers": [
            "Nomor Invoice",
            "Nama Pelanggan",
            "Tanggal Invoice",
            "Jatuh Tempo",
            "Total Harga",
            "Status",
            "Tanggal Bayar",
        ],
        "field_mapping": {
            "Nomor Invoice": "nomor_invoice",
            "Nama Pelanggan": "pelanggan_nama",
            "Tanggal Invoice": "tanggal_invoice",
            "Jatuh Tempo": "tanggal_jatuh_tempo",
            "Total Harga": "total_harga",
            "Status": "status_invoice",
            "Tanggal Bayar": "paid_at",
        },
        "transform_functions": {
            "Total Harga": lambda x: f"Rp {x:,.0f}" if x else "Rp 0",
            "Tanggal Invoice": lambda x: str(x) if x else "",
            "Jatuh Tempo": lambda x: str(x) if x else "",
            "Tanggal Bayar": lambda x: str(x) if x else "",
        },
    }

    LANGGANAN_EXPORT = {
        "headers": [
            "ID",
            "Nama Pelanggan",
            "Email",
            "No Telepon",
            "Alamat",
            "Paket",
            "Harga Paket",
            "Status",
            "Tanggal Aktif",
            "Jatuh Tempo",
            "Brand",
        ],
        "field_mapping": {
            "ID": "id",
            "Nama Pelanggan": "pelanggan_nama",
            "Email": "pelanggan_email",
            "No Telepon": "pelanggan_no_telp",
            "Alamat": "pelanggan_alamat",
            "Paket": "paket_nama",
            "Harga Paket": "paket_harga",
            "Status": "status_langganan",
            "Tanggal Aktif": "tanggal_aktif",
            "Jatuh Tempo": "tanggal_jatuh_tempo",
            "Brand": "brand",
        },
        "transform_functions": {
            "Harga Paket": lambda x: f"Rp {x:,.0f}" if x else "Rp 0",
            "Tanggal Aktif": lambda x: str(x) if x else "",
            "Jatuh Tempo": lambda x: str(x) if x else "",
            "ID": lambda x: str(x) if x else "",
        },
    }


# Factory functions untuk common export patterns
def create_pelanggan_export_response(data: List[Any], export_format: str = "csv") -> StreamingResponse:
    """Factory function untuk pelanggan export"""
    config = ExportConfigurations.PELANGGAN_EXPORT

    processed_data = CSVExporter.prepare_export_data(
        data,
        field_mapping=config["field_mapping"],  # type: ignore
        exclude_fields=config["exclude_fields"],  # type: ignore
        transform_functions=config["transform_functions"],  # type: ignore
    )

    return ExportManager.create_export_response(
        processed_data,
        "pelanggan",
        export_format,
        config["headers"]  # type: ignore
    )


def create_langganan_export_response(data: List[Any], export_format: str = "csv") -> StreamingResponse:
    """Factory function untuk langganan export"""
    config = ExportConfigurations.LANGGANAN_EXPORT

    processed_data = CSVExporter.prepare_export_data(
        data,
        field_mapping=config["field_mapping"],  # type: ignore
        transform_functions=config["transform_functions"],  # type: ignore
    )

    return ExportManager.create_export_response(
        processed_data,
        "langganan",
        export_format,
        config["headers"]  # type: ignore
    )


def create_data_teknis_export_response(data: List[Any], export_format: str = "csv") -> StreamingResponse:
    """Factory function untuk data teknis export"""
    config = ExportConfigurations.DATA_TEKNIS_EXPORT

    processed_data = CSVExporter.prepare_export_data(
        data,
        field_mapping=config["field_mapping"],
        exclude_fields=config["exclude_fields"]  # type: ignore
    )

    return ExportManager.create_export_response(
        processed_data,
        "data_teknis",
        export_format,
        config["headers"]  # type: ignore
    )


def create_invoice_export_response(data: List[Any], export_format: str = "csv") -> StreamingResponse:
    """Factory function untuk invoice export"""
    config = ExportConfigurations.INVOICE_EXPORT

    processed_data = CSVExporter.prepare_export_data(
        data,
        field_mapping=config["field_mapping"],
        transform_functions=config["transform_functions"]  # type: ignore
    )

    return ExportManager.create_export_response(
        processed_data,
        "invoice",
        export_format,
        config["headers"]  # type: ignore
    )